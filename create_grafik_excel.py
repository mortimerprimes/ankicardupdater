#!/usr/bin/env python3
"""
Tagesprotokoll mit Grafiken - Verbesserte Version
- Hormone nur 1x pro Monat (realistisch)
- Ausführliche Erklärungen für den Nutzer
- Schöne, verständliche Grafiken
"""
import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, ScatterChart, BarChart, Reference
from openpyxl.chart.series import XYSeries
from openpyxl.chart.marker import Marker
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_FILE = '/Users/lucasschmidt/Downloads/Tagesprotokoll_mit_Grafiken.xlsx'

np.random.seed(42)
n_days = 30
dates = pd.date_range(start='2026-01-01', periods=n_days, freq='D')

# ============================================
# HORMONE: Nur 2-3x im Monat gemessen (realistisch!)
# ============================================
# Erstelle leere Arrays für Hormone
testosteron_sparse = np.full(n_days, np.nan)
oestradiol_sparse = np.full(n_days, np.nan)

# Nur an Tag 1, 15 und 28 gemessen (realistisch: ca. 1x pro 2 Wochen)
mess_tage = [0, 14, 27]  # Index 0-basiert
testosteron_werte = [520, 480, 510]  # Beispielwerte in ng/dL
oestradiol_werte = [32, 28, 35]  # Beispielwerte in pg/mL

for i, tag in enumerate(mess_tage):
    testosteron_sparse[tag] = testosteron_werte[i]
    oestradiol_sparse[tag] = oestradiol_werte[i]

# Für Korrelationsberechnung: Interpoliere oder nutze nur gemessene Tage
testosteron_for_corr = np.interp(range(n_days), mess_tage, testosteron_werte)
oestradiol_for_corr = np.interp(range(n_days), mess_tage, oestradiol_werte)

def generate_correlated(base, corr_strength, noise=1.5):
    normalized = (base - base.min()) / (base.max() - base.min())
    values = 3 + 5 * normalized * corr_strength + np.random.normal(0, noise, len(base))
    return np.clip(values, 1, 10).round(1)

# Frueh-Werte (stark korreliert)
libido_frueh = generate_correlated(testosteron, 0.8, 1.2)
erektion_frueh = generate_correlated(testosteron, 0.9, 1.0)
energie_frueh = generate_correlated(testosteron, 0.6, 1.5)
stimmung_frueh = generate_correlated(testosteron, 0.5, 1.8)
schlaf_frueh = generate_correlated(oestradiol, -0.3, 2.0)

# Mittag-Werte
libido_mittag = generate_correlated(testosteron, 0.6, 1.5)
erektion_mittag = generate_correlated(testosteron, 0.5, 1.8)
energie_mittag = generate_correlated(testosteron, 0.7, 1.3)
stimmung_mittag = generate_correlated(testosteron, 0.4, 2.0)

# Abend-Werte
libido_abend = generate_correlated(testosteron, 0.7, 1.4) + np.random.uniform(1, 3, n_days)
libido_abend = np.clip(libido_abend, 1, 10).round(1)
erektion_abend = generate_correlated(testosteron, 0.6, 1.6)
energie_abend = generate_correlated(testosteron, 0.4, 2.0)
stimmung_abend = generate_correlated(testosteron, 0.5, 1.7)

# DataFrames
df_frueh = pd.DataFrame({
    'Tag': range(1, n_days+1), 'Datum': dates,
    'Libido': libido_frueh, 'Erektion': erektion_frueh,
    'Energie': energie_frueh, 'Stimmung': stimmung_frueh,
    'Schlaf': schlaf_frueh, 'Testosteron': testosteron.round(0),
    'Oestradiol': oestradiol.round(1)
})

df_mittag = pd.DataFrame({
    'Tag': range(1, n_days+1), 'Datum': dates,
    'Libido': libido_mittag, 'Erektion': erektion_mittag,
    'Energie': energie_mittag, 'Stimmung': stimmung_mittag,
    'Testosteron': testosteron.round(0), 'Oestradiol': oestradiol.round(1)
})

df_abend = pd.DataFrame({
    'Tag': range(1, n_days+1), 'Datum': dates,
    'Libido': libido_abend, 'Erektion': erektion_abend,
    'Energie': energie_abend, 'Stimmung': stimmung_abend,
    'Testosteron': testosteron.round(0), 'Oestradiol': oestradiol.round(1)
})

wb = Workbook()

# === FRUEH Sheet ===
ws_frueh = wb.active
ws_frueh.title = 'Frueh'
for r in dataframe_to_rows(df_frueh, index=False, header=True):
    ws_frueh.append(r)

chart1 = LineChart()
chart1.title = 'FRUEH - Verlauf aller Parameter'
chart1.height = 12
chart1.width = 20
chart1.y_axis.title = "Wert (0-10)"
data = Reference(ws_frueh, min_col=3, min_row=1, max_col=7, max_row=n_days+1)
cats = Reference(ws_frueh, min_col=1, min_row=2, max_row=n_days+1)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws_frueh.add_chart(chart1, 'K2')

chart2 = LineChart()
chart2.title = 'Testosteron-Verlauf'
chart2.height = 10
chart2.width = 20
chart2.y_axis.title = "ng/dL"
data2 = Reference(ws_frueh, min_col=8, min_row=1, max_col=8, max_row=n_days+1)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats)
ws_frueh.add_chart(chart2, 'K18')

# === MITTAG Sheet ===
ws_mittag = wb.create_sheet('Mittag')
for r in dataframe_to_rows(df_mittag, index=False, header=True):
    ws_mittag.append(r)

chart3 = LineChart()
chart3.title = 'MITTAG - Verlauf aller Parameter'
chart3.height = 12
chart3.width = 20
chart3.y_axis.title = "Wert (0-10)"
data3 = Reference(ws_mittag, min_col=3, min_row=1, max_col=6, max_row=n_days+1)
cats3 = Reference(ws_mittag, min_col=1, min_row=2, max_row=n_days+1)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws_mittag.add_chart(chart3, 'J2')

# === ABEND Sheet ===
ws_abend = wb.create_sheet('Abend')
for r in dataframe_to_rows(df_abend, index=False, header=True):
    ws_abend.append(r)

chart4 = LineChart()
chart4.title = 'ABEND - Verlauf aller Parameter'
chart4.height = 12
chart4.width = 20
chart4.y_axis.title = "Wert (0-10)"
data4 = Reference(ws_abend, min_col=3, min_row=1, max_col=6, max_row=n_days+1)
cats4 = Reference(ws_abend, min_col=1, min_row=2, max_row=n_days+1)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
ws_abend.add_chart(chart4, 'J2')

# === Vergleich Sheet ===
ws_vgl = wb.create_sheet('Vergleich_Zeitpunkte')
vgl_data = pd.DataFrame({
    'Parameter': ['Libido', 'Erektion', 'Energie', 'Stimmung'],
    'Frueh': [df_frueh['Libido'].mean(), df_frueh['Erektion'].mean(), df_frueh['Energie'].mean(), df_frueh['Stimmung'].mean()],
    'Mittag': [df_mittag['Libido'].mean(), df_mittag['Erektion'].mean(), df_mittag['Energie'].mean(), df_mittag['Stimmung'].mean()],
    'Abend': [df_abend['Libido'].mean(), df_abend['Erektion'].mean(), df_abend['Energie'].mean(), df_abend['Stimmung'].mean()]
}).round(2)
for r in dataframe_to_rows(vgl_data, index=False, header=True):
    ws_vgl.append(r)

chart5 = BarChart()
chart5.type = 'col'
chart5.title = 'Durchschnittswerte: Frueh vs Mittag vs Abend'
chart5.height = 14
chart5.width = 18
chart5.y_axis.title = "Durchschnitt (0-10)"
data5 = Reference(ws_vgl, min_col=2, min_row=1, max_col=4, max_row=5)
cats5 = Reference(ws_vgl, min_col=1, min_row=2, max_row=5)
chart5.add_data(data5, titles_from_data=True)
chart5.set_categories(cats5)
ws_vgl.add_chart(chart5, 'F2')

# === Korrelation Testosteron ===
ws_corr = wb.create_sheet('Korr_Testosteron')
corr_data = pd.DataFrame({
    'Parameter': ['Libido', 'Erektion', 'Energie', 'Stimmung'],
    'Frueh': [
        round(stats.pearsonr(df_frueh['Testosteron'], df_frueh['Libido'])[0], 3),
        round(stats.pearsonr(df_frueh['Testosteron'], df_frueh['Erektion'])[0], 3),
        round(stats.pearsonr(df_frueh['Testosteron'], df_frueh['Energie'])[0], 3),
        round(stats.pearsonr(df_frueh['Testosteron'], df_frueh['Stimmung'])[0], 3)
    ],
    'Mittag': [
        round(stats.pearsonr(df_mittag['Testosteron'], df_mittag['Libido'])[0], 3),
        round(stats.pearsonr(df_mittag['Testosteron'], df_mittag['Erektion'])[0], 3),
        round(stats.pearsonr(df_mittag['Testosteron'], df_mittag['Energie'])[0], 3),
        round(stats.pearsonr(df_mittag['Testosteron'], df_mittag['Stimmung'])[0], 3)
    ],
    'Abend': [
        round(stats.pearsonr(df_abend['Testosteron'], df_abend['Libido'])[0], 3),
        round(stats.pearsonr(df_abend['Testosteron'], df_abend['Erektion'])[0], 3),
        round(stats.pearsonr(df_abend['Testosteron'], df_abend['Energie'])[0], 3),
        round(stats.pearsonr(df_abend['Testosteron'], df_abend['Stimmung'])[0], 3)
    ]
})
for r in dataframe_to_rows(corr_data, index=False, header=True):
    ws_corr.append(r)

ws_corr['A7'] = 'INTERPRETATION:'
ws_corr['A8'] = 'r > 0.5: Starke positive Korrelation'
ws_corr['A9'] = 'r 0.3-0.5: Moderate Korrelation'
ws_corr['A10'] = 'r < 0.3: Schwache Korrelation'
ws_corr['A11'] = 'r < 0: Negative Korrelation'

chart6 = BarChart()
chart6.type = 'col'
chart6.title = 'TESTOSTERON-Korrelationen nach Tageszeit'
chart6.height = 14
chart6.width = 20
chart6.y_axis.title = "Korrelation (r)"
data6 = Reference(ws_corr, min_col=2, min_row=1, max_col=4, max_row=5)
cats6 = Reference(ws_corr, min_col=1, min_row=2, max_row=5)
chart6.add_data(data6, titles_from_data=True)
chart6.set_categories(cats6)
ws_corr.add_chart(chart6, 'F2')

# === Scatter Plots ===
ws_scatter = wb.create_sheet('Scatter_Testosteron')
scatter_df = pd.DataFrame({
    'Testosteron': testosteron.round(0),
    'Libido_Frueh': libido_frueh,
    'Erektion_Frueh': erektion_frueh,
    'Libido_Abend': libido_abend
})
for r in dataframe_to_rows(scatter_df, index=False, header=True):
    ws_scatter.append(r)

# Scatter 1: Testosteron vs Libido Frueh
sc1 = ScatterChart()
sc1.title = 'Testosteron vs Libido (Frueh)'
sc1.x_axis.title = "Testosteron (ng/dL)"
sc1.y_axis.title = "Libido (0-10)"
sc1.height = 12
sc1.width = 14
xv = Reference(ws_scatter, min_col=1, min_row=2, max_row=n_days+1)
yv = Reference(ws_scatter, min_col=2, min_row=2, max_row=n_days+1)
s1 = Series(yv, xv, title='Libido Frueh')
sc1.series.append(s1)
ws_scatter.add_chart(sc1, 'F2')

# Scatter 2: Testosteron vs Erektion Frueh
sc2 = ScatterChart()
sc2.title = 'Testosteron vs Erektion (Frueh)'
sc2.x_axis.title = "Testosteron (ng/dL)"
sc2.y_axis.title = "Erektion (0-10)"
sc2.height = 12
sc2.width = 14
yv2 = Reference(ws_scatter, min_col=3, min_row=2, max_row=n_days+1)
s2 = Series(yv2, xv, title='Erektion Frueh')
sc2.series.append(s2)
ws_scatter.add_chart(sc2, 'F18')

# Scatter 3: Testosteron vs Libido Abend
sc3 = ScatterChart()
sc3.title = 'Testosteron vs Libido (Abend)'
sc3.x_axis.title = "Testosteron (ng/dL)"
sc3.y_axis.title = "Libido (0-10)"
sc3.height = 12
sc3.width = 14
yv3 = Reference(ws_scatter, min_col=4, min_row=2, max_row=n_days+1)
s3 = Series(yv3, xv, title='Libido Abend')
sc3.series.append(s3)
ws_scatter.add_chart(sc3, 'T2')

# === Oestradiol Korrelation ===
ws_oe = wb.create_sheet('Korr_Oestradiol')
oe_data = pd.DataFrame({
    'Parameter': ['Libido', 'Erektion', 'Energie', 'Stimmung', 'Schlaf'],
    'Korrelation_r': [
        round(stats.pearsonr(df_frueh['Oestradiol'], df_frueh['Libido'])[0], 3),
        round(stats.pearsonr(df_frueh['Oestradiol'], df_frueh['Erektion'])[0], 3),
        round(stats.pearsonr(df_frueh['Oestradiol'], df_frueh['Energie'])[0], 3),
        round(stats.pearsonr(df_frueh['Oestradiol'], df_frueh['Stimmung'])[0], 3),
        round(stats.pearsonr(df_frueh['Oestradiol'], df_frueh['Schlaf'])[0], 3)
    ]
})
for r in dataframe_to_rows(oe_data, index=False, header=True):
    ws_oe.append(r)

chart7 = BarChart()
chart7.type = 'col'
chart7.title = 'OESTRADIOL-Korrelationen (Frueh)'
chart7.height = 14
chart7.width = 16
chart7.y_axis.title = "Korrelation (r)"
data7 = Reference(ws_oe, min_col=2, min_row=1, max_col=2, max_row=6)
cats7 = Reference(ws_oe, min_col=1, min_row=2, max_row=6)
chart7.add_data(data7, titles_from_data=True)
chart7.set_categories(cats7)
ws_oe.add_chart(chart7, 'D2')

# === Dashboard ===
ws_d = wb.create_sheet('Dashboard', 0)
ws_d['A1'] = 'TAGESPROTOKOLL - GRAFISCHES DASHBOARD'
ws_d['A1'].font = Font(bold=True, size=16)

ws_d['A3'] = 'ZUSAMMENFASSUNG (30 Tage Testdaten)'
ws_d['A3'].font = Font(bold=True, size=12)

ws_d['A5'] = 'Durchschnittswerte:'
ws_d['B5'] = 'Frueh'
ws_d['C5'] = 'Mittag'
ws_d['D5'] = 'Abend'

ws_d['A6'] = 'Libido'
ws_d['B6'] = round(df_frueh['Libido'].mean(), 1)
ws_d['C6'] = round(df_mittag['Libido'].mean(), 1)
ws_d['D6'] = round(df_abend['Libido'].mean(), 1)

ws_d['A7'] = 'Erektion'
ws_d['B7'] = round(df_frueh['Erektion'].mean(), 1)
ws_d['C7'] = round(df_mittag['Erektion'].mean(), 1)
ws_d['D7'] = round(df_abend['Erektion'].mean(), 1)

ws_d['A8'] = 'Energie'
ws_d['B8'] = round(df_frueh['Energie'].mean(), 1)
ws_d['C8'] = round(df_mittag['Energie'].mean(), 1)
ws_d['D8'] = round(df_abend['Energie'].mean(), 1)

ws_d['A9'] = 'Stimmung'
ws_d['B9'] = round(df_frueh['Stimmung'].mean(), 1)
ws_d['C9'] = round(df_mittag['Stimmung'].mean(), 1)
ws_d['D9'] = round(df_abend['Stimmung'].mean(), 1)

ws_d['A11'] = 'Hormone:'
ws_d['A12'] = 'Testosteron (Durchschnitt):'
ws_d['B12'] = f'{round(testosteron.mean(), 0)} ng/dL'
ws_d['A13'] = 'Oestradiol (Durchschnitt):'
ws_d['B13'] = f'{round(oestradiol.mean(), 1)} pg/mL'

ws_d['A15'] = 'STAERKSTE TESTOSTERON-KORRELATIONEN:'
ws_d['A15'].font = Font(bold=True)
ws_d['A16'] = f'1. Erektion (Frueh): r = {round(stats.pearsonr(df_frueh["Testosteron"], df_frueh["Erektion"])[0], 2)}'
ws_d['A17'] = f'2. Libido (Frueh): r = {round(stats.pearsonr(df_frueh["Testosteron"], df_frueh["Libido"])[0], 2)}'
ws_d['A18'] = f'3. Energie (Mittag): r = {round(stats.pearsonr(df_mittag["Testosteron"], df_mittag["Energie"])[0], 2)}'

ws_d['A20'] = '>>> Hoeheres Testosteron korreliert am staerksten mit Erektion und Libido am Morgen!'
ws_d['A20'].font = Font(italic=True)

ws_d['A22'] = 'SHEETS MIT GRAFIKEN:'
ws_d['A23'] = '- Frueh: Liniendiagramm aller Parameter + Testosteron'
ws_d['A24'] = '- Mittag: Liniendiagramm aller Parameter'
ws_d['A25'] = '- Abend: Liniendiagramm aller Parameter'
ws_d['A26'] = '- Vergleich_Zeitpunkte: Balkendiagramm Frueh vs Mittag vs Abend'
ws_d['A27'] = '- Korr_Testosteron: Balkendiagramm der Korrelationen'
ws_d['A28'] = '- Scatter_Testosteron: Streudiagramme (zeigen Zusammenhang)'
ws_d['A29'] = '- Korr_Oestradiol: Oestradiol-Einfluss auf Parameter'

# Spaltenbreiten
for ws in wb.worksheets:
    for col in ws.columns:
        try:
            ws.column_dimensions[col[0].column_letter].width = 14
        except:
            pass

wb.save(OUTPUT_FILE)
print(f'Datei erstellt: {OUTPUT_FILE}')
print('\nEnthaltene Sheets mit Grafiken:')
print('  1. Dashboard - Uebersicht')
print('  2. Frueh - Liniendiagramme')
print('  3. Mittag - Liniendiagramme')
print('  4. Abend - Liniendiagramme')
print('  5. Vergleich_Zeitpunkte - Balkendiagramm')
print('  6. Korr_Testosteron - Korrelations-Balkendiagramm')
print('  7. Scatter_Testosteron - 3 Streudiagramme')
print('  8. Korr_Oestradiol - Oestradiol-Korrelationen')

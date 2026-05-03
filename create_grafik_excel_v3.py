#!/usr/bin/env python3
"""
Tagesprotokoll - Finale Version
- Östradiol UND Testosteron Einfluss
- Tagesdurchschnitte mit Kurven
- Komplett durchdacht und benutzerfreundlich
"""
import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_FILE = '/Users/lucasschmidt/Downloads/Tagesprotokoll_mit_Grafiken.xlsx'

# ============================================
# STYLING
# ============================================
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16)
SUBTITLE_FONT = Font(bold=True, size=12)
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

np.random.seed(42)
n_days = 30
dates = pd.date_range(start='2026-01-01', periods=n_days, freq='D')

# ============================================
# HORMONE (nur 3 Messungen im Monat - realistisch)
# ============================================
mess_tage = [0, 14, 27]  # Tag 1, 15, 28
testosteron_werte = [520, 450, 490]  # ng/dL
oestradiol_werte = [32, 38, 28]  # pg/mL - variiert gegenläufig

# Für Berechnungen: Forward-Fill
testo_sparse = np.full(n_days, np.nan)
oestra_sparse = np.full(n_days, np.nan)
for i, tag in enumerate(mess_tage):
    testo_sparse[tag] = testosteron_werte[i]
    oestra_sparse[tag] = oestradiol_werte[i]

testo_filled = pd.Series(testo_sparse).ffill().bfill().values
oestra_filled = pd.Series(oestra_sparse).ffill().bfill().values

# ============================================
# TÄGLICHE WERTE GENERIEREN
# ============================================
def gen_values(base, strength, noise=1.5):
    """Generiert korrelierte Werte"""
    norm = (base - base.min()) / (base.max() - base.min() + 0.001)
    vals = 4 + 4 * norm * strength + np.random.normal(0, noise, len(base))
    return np.clip(vals, 1, 10).round(1)

# FRÜH
libido_f = gen_values(testo_filled, 0.7, 1.0)
erektion_f = gen_values(testo_filled, 0.8, 0.9)
energie_f = gen_values(testo_filled, 0.5, 1.2)
stimmung_f = gen_values(testo_filled, 0.4, 1.5)
schlaf = np.random.uniform(5, 9, n_days).round(1)

# MITTAG  
libido_m = gen_values(testo_filled, 0.5, 1.2)
erektion_m = gen_values(testo_filled, 0.4, 1.4)
energie_m = gen_values(testo_filled, 0.6, 1.0)
stimmung_m = gen_values(testo_filled, 0.3, 1.6)

# ABEND (Libido typischerweise höher)
libido_a = np.clip(gen_values(testo_filled, 0.6, 1.1) + 1.0, 1, 10).round(1)
erektion_a = gen_values(testo_filled, 0.5, 1.3)
energie_a = gen_values(testo_filled, 0.3, 1.4)
stimmung_a = gen_values(testo_filled, 0.4, 1.3)

# Östradiol-Effekt: Hohe Werte -> leicht negativer Effekt auf Erektion
erektion_f = np.clip(erektion_f - 0.3 * (oestra_filled - 30) / 10, 1, 10).round(1)
erektion_a = np.clip(erektion_a - 0.3 * (oestra_filled - 30) / 10, 1, 10).round(1)

# ============================================
# TAGESDURCHSCHNITTE BERECHNEN
# ============================================
libido_avg = ((libido_f + libido_m + libido_a) / 3).round(1)
erektion_avg = ((erektion_f + erektion_m + erektion_a) / 3).round(1)
energie_avg = ((energie_f + energie_m + energie_a) / 3).round(1)
stimmung_avg = ((stimmung_f + stimmung_m + stimmung_a) / 3).round(1)

# ============================================
# WORKBOOK ERSTELLEN
# ============================================
wb = Workbook()

# ============================================
# 1. ANLEITUNG
# ============================================
ws = wb.active
ws.title = '1_Anleitung'

content = [
    ['TAGESPROTOKOLL - BEDIENUNGSANLEITUNG'],
    [''],
    ['SO NUTZT DU DIESES DOKUMENT:'],
    [''],
    ['SCHRITT 1: Tägliche Werte eintragen'],
    ['→ Gehe zu den Sheets "2_Früh", "3_Mittag", "4_Abend"'],
    ['→ Trage jeden Tag deine Werte ein (Skala 0-10)'],
    ['→ 0 = sehr schlecht, 10 = ausgezeichnet'],
    [''],
    ['SCHRITT 2: Hormonwerte eintragen (bei Blutabnahme)'],
    ['→ Gehe zu Sheet "5_Hormonwerte"'],
    ['→ Trage Testosteron und Östradiol ein (nur wenn gemessen!)'],
    [''],
    ['SCHRITT 3: Auswertung ansehen'],
    ['→ "6_Tagesdurchschnitt" zeigt deinen Verlauf über den Monat'],
    ['→ "7_Testosteron_Einfluss" zeigt wie Testosteron dich beeinflusst'],
    ['→ "8_Östradiol_Einfluss" zeigt wie Östradiol dich beeinflusst'],
    ['→ "9_Zusammenfassung" gibt dir den Überblick'],
    [''],
    ['─────────────────────────────────────────────────'],
    [''],
    ['WAS BEDEUTEN DIE KORRELATIONEN?'],
    [''],
    ['Eine Korrelation (r) zeigt den Zusammenhang zwischen zwei Werten:'],
    [''],
    ['   r = +0.7 bis +1.0  →  STARKER POSITIVER Zusammenhang'],
    ['   Beispiel: Höheres Testosteron = Höhere Libido'],
    [''],
    ['   r = +0.3 bis +0.7  →  MODERATER POSITIVER Zusammenhang'],
    ['   Beispiel: Testosteron beeinflusst den Wert etwas'],
    [''],
    ['   r = -0.3 bis +0.3  →  KEIN deutlicher Zusammenhang'],
    ['   Beispiel: Hormon hat keinen Einfluss auf diesen Wert'],
    [''],
    ['   r = -0.7 bis -0.3  →  MODERATER NEGATIVER Zusammenhang'],
    ['   Beispiel: Höheres Hormon = Niedrigerer Wert'],
    [''],
    ['   r = -1.0 bis -0.7  →  STARKER NEGATIVER Zusammenhang'],
    [''],
    ['─────────────────────────────────────────────────'],
    [''],
    ['FARBCODE IN DEN TABELLEN:'],
    ['   Grün  = Positiver Zusammenhang (Hormon hilft)'],
    ['   Gelb  = Kein/schwacher Zusammenhang'],
    ['   Rot   = Negativer Zusammenhang (Hormon schadet)'],
    [''],
    ['─────────────────────────────────────────────────'],
    [''],
    ['TIPPS:'],
    ['• Sammle mindestens 2-3 Wochen Daten für aussagekräftige Ergebnisse'],
    ['• Trage Werte möglichst zur gleichen Zeit ein'],
    ['• Sei ehrlich - nur so helfen dir die Daten wirklich'],
]

for i, row in enumerate(content, 1):
    cell = ws.cell(row=i, column=1)
    cell.value = row[0] if row else ''
    if i == 1:
        cell.font = TITLE_FONT
    elif row and (row[0].startswith('SCHRITT') or row[0].startswith('WAS BEDEUTEN') or 
                  row[0].startswith('FARBCODE') or row[0].startswith('TIPPS') or 
                  row[0].startswith('SO NUTZT')):
        cell.font = SUBTITLE_FONT

ws.column_dimensions['A'].width = 70

# ============================================
# 2. FRÜH
# ============================================
ws_f = wb.create_sheet('2_Früh')

df_f = pd.DataFrame({
    'Tag': range(1, n_days+1),
    'Datum': dates.strftime('%d.%m'),
    'Libido': libido_f,
    'Erektion': erektion_f,
    'Energie': energie_f,
    'Stimmung': stimmung_f,
    'Schlaf': schlaf
})

for r in dataframe_to_rows(df_f, index=False, header=True):
    ws_f.append(r)

# Header formatieren
for col in range(1, 8):
    ws_f.cell(row=1, column=col).fill = HEADER_FILL
    ws_f.cell(row=1, column=col).font = HEADER_FONT

# Chart
chart = LineChart()
chart.title = "MORGENS - Deine Werte im Verlauf"
chart.style = 10
chart.height = 12
chart.width = 16
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 10
chart.y_axis.title = "Wert (0-10)"
chart.x_axis.title = "Tag"

data = Reference(ws_f, min_col=3, min_row=1, max_col=7, max_row=n_days+1)
cats = Reference(ws_f, min_col=1, min_row=2, max_row=n_days+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_f.add_chart(chart, 'I2')

for col in ['A','B','C','D','E','F','G']:
    ws_f.column_dimensions[col].width = 10

# ============================================
# 3. MITTAG
# ============================================
ws_m = wb.create_sheet('3_Mittag')

df_m = pd.DataFrame({
    'Tag': range(1, n_days+1),
    'Datum': dates.strftime('%d.%m'),
    'Libido': libido_m,
    'Erektion': erektion_m,
    'Energie': energie_m,
    'Stimmung': stimmung_m
})

for r in dataframe_to_rows(df_m, index=False, header=True):
    ws_m.append(r)

for col in range(1, 7):
    ws_m.cell(row=1, column=col).fill = HEADER_FILL
    ws_m.cell(row=1, column=col).font = HEADER_FONT

chart = LineChart()
chart.title = "MITTAGS - Deine Werte im Verlauf"
chart.style = 10
chart.height = 12
chart.width = 16
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 10

data = Reference(ws_m, min_col=3, min_row=1, max_col=6, max_row=n_days+1)
cats = Reference(ws_m, min_col=1, min_row=2, max_row=n_days+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_m.add_chart(chart, 'H2')

# ============================================
# 4. ABEND
# ============================================
ws_a = wb.create_sheet('4_Abend')

df_a = pd.DataFrame({
    'Tag': range(1, n_days+1),
    'Datum': dates.strftime('%d.%m'),
    'Libido': libido_a,
    'Erektion': erektion_a,
    'Energie': energie_a,
    'Stimmung': stimmung_a
})

for r in dataframe_to_rows(df_a, index=False, header=True):
    ws_a.append(r)

for col in range(1, 7):
    ws_a.cell(row=1, column=col).fill = HEADER_FILL
    ws_a.cell(row=1, column=col).font = HEADER_FONT

chart = LineChart()
chart.title = "ABENDS - Deine Werte im Verlauf"
chart.style = 10
chart.height = 12
chart.width = 16
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 10

data = Reference(ws_a, min_col=3, min_row=1, max_col=6, max_row=n_days+1)
cats = Reference(ws_a, min_col=1, min_row=2, max_row=n_days+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_a.add_chart(chart, 'H2')

# ============================================
# 5. HORMONWERTE
# ============================================
ws_h = wb.create_sheet('5_Hormonwerte')

ws_h['A1'] = 'DEINE HORMONWERTE'
ws_h['A1'].font = TITLE_FONT

ws_h['A3'] = 'Trage hier deine Laborwerte ein (nur bei Blutabnahme!):'
ws_h['A3'].font = Font(italic=True)

headers = ['Datum', 'Tag', 'Testosteron (ng/dL)', 'Östradiol (pg/mL)', 'Notizen']
for i, h in enumerate(headers, 1):
    cell = ws_h.cell(row=5, column=i)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

# Beispieldaten eintragen
for idx, tag in enumerate(mess_tage):
    row = 6 + idx
    ws_h.cell(row=row, column=1).value = dates[tag].strftime('%d.%m.%Y')
    ws_h.cell(row=row, column=2).value = tag + 1
    ws_h.cell(row=row, column=3).value = testosteron_werte[idx]
    ws_h.cell(row=row, column=4).value = oestradiol_werte[idx]

# Referenzbereiche
ws_h['A11'] = 'REFERENZBEREICHE:'
ws_h['A11'].font = SUBTITLE_FONT
ws_h['A12'] = 'Testosteron (Männer): 300-1000 ng/dL (optimal: 500-800)'
ws_h['A13'] = 'Östradiol (Männer): 10-40 pg/mL (optimal: 20-30)'
ws_h['A14'] = ''
ws_h['A15'] = 'WICHTIG:'
ws_h['A15'].font = SUBTITLE_FONT
ws_h['A16'] = '• Zu hohes Östradiol kann Libido und Erektion verschlechtern'
ws_h['A17'] = '• Optimales Verhältnis Testosteron:Östradiol ca. 15-20:1'

for col in ['A','B','C','D','E']:
    ws_h.column_dimensions[col].width = 20

# ============================================
# 6. TAGESDURCHSCHNITT (NEU!)
# ============================================
ws_avg = wb.create_sheet('6_Tagesdurchschnitt')

ws_avg['A1'] = 'TAGESDURCHSCHNITT - Dein Gesamtverlauf'
ws_avg['A1'].font = TITLE_FONT
ws_avg['A2'] = '(Durchschnitt aus Früh + Mittag + Abend)'
ws_avg['A2'].font = Font(italic=True, color="666666")

# Tabelle
headers = ['Tag', 'Datum', 'Ø Libido', 'Ø Erektion', 'Ø Energie', 'Ø Stimmung', 'Schlaf']
for i, h in enumerate(headers, 1):
    cell = ws_avg.cell(row=4, column=i)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

for day in range(n_days):
    row = 5 + day
    ws_avg.cell(row=row, column=1).value = day + 1
    ws_avg.cell(row=row, column=2).value = dates[day].strftime('%d.%m')
    ws_avg.cell(row=row, column=3).value = libido_avg[day]
    ws_avg.cell(row=row, column=4).value = erektion_avg[day]
    ws_avg.cell(row=row, column=5).value = energie_avg[day]
    ws_avg.cell(row=row, column=6).value = stimmung_avg[day]
    ws_avg.cell(row=row, column=7).value = schlaf[day]

# Chart für Tagesdurchschnitt
chart = LineChart()
chart.title = "TAGESDURCHSCHNITT - So geht es dir im Verlauf"
chart.style = 10
chart.height = 14
chart.width = 18
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 10
chart.y_axis.title = "Durchschnittswert (0-10)"
chart.x_axis.title = "Tag"

data = Reference(ws_avg, min_col=3, min_row=4, max_col=7, max_row=4+n_days)
cats = Reference(ws_avg, min_col=1, min_row=5, max_row=4+n_days)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_avg.add_chart(chart, 'I4')

# Erklärung
ws_avg.cell(row=n_days+7, column=1).value = 'WAS ZEIGT DIESE GRAFIK?'
ws_avg.cell(row=n_days+7, column=1).font = SUBTITLE_FONT
ws_avg.cell(row=n_days+8, column=1).value = 'Der Tagesdurchschnitt fasst deine Früh-, Mittag- und Abendwerte zusammen.'
ws_avg.cell(row=n_days+9, column=1).value = 'So siehst du auf einen Blick, wie es dir an jedem Tag insgesamt ging.'
ws_avg.cell(row=n_days+10, column=1).value = 'Vergleiche mit den Hormonmessungen: Fallen Tiefs mit niedrigem Testosteron zusammen?'

for col in ['A','B','C','D','E','F','G']:
    ws_avg.column_dimensions[col].width = 12

# ============================================
# 7. TESTOSTERON-EINFLUSS
# ============================================
ws_t = wb.create_sheet('7_Testosteron_Einfluss')

ws_t['A1'] = 'WIE BEEINFLUSST TESTOSTERON DEINE WERTE?'
ws_t['A1'].font = TITLE_FONT

# Korrelationen berechnen
def calc_corr(x, y):
    try:
        r, p = stats.pearsonr(x, y)
        return round(r, 2), round(p, 3)
    except:
        return 0, 1

def interpret(r):
    if r >= 0.5: return "Stark positiv"
    elif r >= 0.3: return "Moderat positiv"
    elif r >= -0.3: return "Kein Einfluss"
    elif r >= -0.5: return "Moderat negativ"
    else: return "Stark negativ"

def meaning(r, param):
    if r >= 0.3:
        return f"Höheres Testosteron → bessere {param}"
    elif r <= -0.3:
        return f"Höheres Testosteron → schlechtere {param}"
    else:
        return f"Testosteron beeinflusst {param} kaum"

# Header
headers = ['Parameter', 'Korrelation (r)', 'Stärke', 'Das bedeutet für dich']
for i, h in enumerate(headers, 1):
    cell = ws_t.cell(row=3, column=i)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

# Korrelationen mit Tagesdurchschnitt (aussagekräftiger)
params = [
    ('Libido (Ø Tag)', libido_avg),
    ('Erektion (Ø Tag)', erektion_avg),
    ('Energie (Ø Tag)', energie_avg),
    ('Stimmung (Ø Tag)', stimmung_avg),
    ('Schlaf', schlaf),
]

row = 4
for name, values in params:
    r, p = calc_corr(testo_filled, values)
    ws_t.cell(row=row, column=1).value = name
    ws_t.cell(row=row, column=2).value = r
    ws_t.cell(row=row, column=3).value = interpret(r)
    ws_t.cell(row=row, column=4).value = meaning(r, name.split(' ')[0])
    
    # Farbcodierung
    if r >= 0.3:
        ws_t.cell(row=row, column=2).fill = GOOD_FILL
    elif r <= -0.3:
        ws_t.cell(row=row, column=2).fill = BAD_FILL
    else:
        ws_t.cell(row=row, column=2).fill = NEUTRAL_FILL
    row += 1

# Balkendiagramm
chart = BarChart()
chart.type = 'bar'
chart.title = "Testosteron-Einfluss auf deine Werte"
chart.style = 10
chart.height = 10
chart.width = 12
chart.x_axis.scaling.min = -1
chart.x_axis.scaling.max = 1
chart.x_axis.title = "Korrelation (r)"

data = Reference(ws_t, min_col=2, min_row=3, max_row=8)
cats = Reference(ws_t, min_col=1, min_row=4, max_row=8)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_t.add_chart(chart, 'F3')

# Erklärung
ws_t['A11'] = 'SO LIEST DU DIE TABELLE:'
ws_t['A11'].font = SUBTITLE_FONT
ws_t['A12'] = '• Grün = Testosteron hat POSITIVEN Einfluss (höher ist besser für dich)'
ws_t['A13'] = '• Gelb = Testosteron hat KEINEN deutlichen Einfluss'
ws_t['A14'] = '• Rot = Testosteron hat NEGATIVEN Einfluss (selten bei Testosteron)'

for col in ['A','B','C','D']:
    ws_t.column_dimensions[col].width = 25

# ============================================
# 8. ÖSTRADIOL-EINFLUSS (NEU!)
# ============================================
ws_e = wb.create_sheet('8_Östradiol_Einfluss')

ws_e['A1'] = 'WIE BEEINFLUSST ÖSTRADIOL DEINE WERTE?'
ws_e['A1'].font = TITLE_FONT

ws_e['A2'] = '(Östradiol ist ein weibliches Hormon, das auch Männer haben - zu viel kann problematisch sein)'
ws_e['A2'].font = Font(italic=True, size=10, color="666666")

# Header
for i, h in enumerate(headers, 1):
    cell = ws_e.cell(row=4, column=i)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

row = 5
for name, values in params:
    r, p = calc_corr(oestra_filled, values)
    ws_e.cell(row=row, column=1).value = name
    ws_e.cell(row=row, column=2).value = r
    ws_e.cell(row=row, column=3).value = interpret(r)
    
    # Bei Östradiol ist NEGATIV oft GUT (weniger Östradiol = besser für Männer)
    if r <= -0.3:
        meaning_text = f"Höheres Östradiol → schlechtere {name.split(' ')[0]} (typisch!)"
        ws_e.cell(row=row, column=2).fill = BAD_FILL
    elif r >= 0.3:
        meaning_text = f"Höheres Östradiol → bessere {name.split(' ')[0]} (ungewöhnlich)"
        ws_e.cell(row=row, column=2).fill = NEUTRAL_FILL
    else:
        meaning_text = f"Östradiol beeinflusst {name.split(' ')[0]} kaum"
        ws_e.cell(row=row, column=2).fill = NEUTRAL_FILL
    
    ws_e.cell(row=row, column=4).value = meaning_text
    row += 1

# Balkendiagramm
chart = BarChart()
chart.type = 'bar'
chart.title = "Östradiol-Einfluss auf deine Werte"
chart.style = 10
chart.height = 10
chart.width = 12
chart.x_axis.scaling.min = -1
chart.x_axis.scaling.max = 1
chart.x_axis.title = "Korrelation (r)"

data = Reference(ws_e, min_col=2, min_row=4, max_row=9)
cats = Reference(ws_e, min_col=1, min_row=5, max_row=9)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_e.add_chart(chart, 'F4')

# Erklärung
ws_e['A12'] = 'WICHTIG ZU WISSEN:'
ws_e['A12'].font = SUBTITLE_FONT
ws_e['A13'] = '• Bei Männern ist NIEDRIGES Östradiol oft besser'
ws_e['A14'] = '• Zu hohes Östradiol (>40 pg/mL) kann Erektion und Libido verschlechtern'
ws_e['A15'] = '• Rote Werte hier bedeuten: Wenn Östradiol steigt, wird der Wert SCHLECHTER'
ws_e['A16'] = '• Das ist bei Männern normal und zu erwarten'
ws_e['A17'] = ''
ws_e['A18'] = 'TIPP:'
ws_e['A18'].font = SUBTITLE_FONT
ws_e['A19'] = 'Wenn du hohe Östradiol-Werte hast und Probleme mit Libido/Erektion,'
ws_e['A20'] = 'sprich mit deinem Arzt über mögliche Ursachen (z.B. Körperfett, Medikamente).'

for col in ['A','B','C','D']:
    ws_e.column_dimensions[col].width = 25

# ============================================
# 9. ZUSAMMENFASSUNG
# ============================================
ws_s = wb.create_sheet('9_Zusammenfassung')

ws_s['A1'] = 'DEINE ZUSAMMENFASSUNG'
ws_s['A1'].font = TITLE_FONT

# Durchschnitte
ws_s['A3'] = 'DEINE DURCHSCHNITTSWERTE (30 Tage):'
ws_s['A3'].font = SUBTITLE_FONT

avg_headers = ['', 'Früh', 'Mittag', 'Abend', 'Tages-Ø']
for i, h in enumerate(avg_headers, 1):
    cell = ws_s.cell(row=5, column=i)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

data_rows = [
    ['Libido', round(libido_f.mean(),1), round(libido_m.mean(),1), round(libido_a.mean(),1), round(libido_avg.mean(),1)],
    ['Erektion', round(erektion_f.mean(),1), round(erektion_m.mean(),1), round(erektion_a.mean(),1), round(erektion_avg.mean(),1)],
    ['Energie', round(energie_f.mean(),1), round(energie_m.mean(),1), round(energie_a.mean(),1), round(energie_avg.mean(),1)],
    ['Stimmung', round(stimmung_f.mean(),1), round(stimmung_m.mean(),1), round(stimmung_a.mean(),1), round(stimmung_avg.mean(),1)],
    ['Schlaf', round(schlaf.mean(),1), '-', '-', '-'],
]

for i, row_data in enumerate(data_rows):
    for j, val in enumerate(row_data):
        ws_s.cell(row=6+i, column=1+j).value = val

# Hormonwerte
ws_s['A13'] = 'DEINE HORMONWERTE:'
ws_s['A13'].font = SUBTITLE_FONT
ws_s['A14'] = f'Testosteron (Ø): {round(np.mean(testosteron_werte))} ng/dL'
ws_s['A15'] = f'Östradiol (Ø): {round(np.mean(oestradiol_werte), 1)} pg/mL'
ws_s['A16'] = f'Verhältnis T:E2: {round(np.mean(testosteron_werte) / np.mean(oestradiol_werte), 1)}:1'

# Erkenntnisse
ws_s['A19'] = 'WICHTIGSTE ERKENNTNISSE:'
ws_s['A19'].font = SUBTITLE_FONT

# Top Testosteron-Korrelation
testo_corrs = [(n, calc_corr(testo_filled, v)[0]) for n, v in params]
best_testo = max(testo_corrs, key=lambda x: abs(x[1]))
ws_s['A20'] = f'• Testosteron beeinflusst am stärksten: {best_testo[0]} (r={best_testo[1]})'

# Bester Zeitpunkt
avg_by_time = {
    'Früh': (libido_f.mean() + erektion_f.mean() + energie_f.mean() + stimmung_f.mean()) / 4,
    'Mittag': (libido_m.mean() + erektion_m.mean() + energie_m.mean() + stimmung_m.mean()) / 4,
    'Abend': (libido_a.mean() + erektion_a.mean() + energie_a.mean() + stimmung_a.mean()) / 4,
}
best_time = max(avg_by_time, key=avg_by_time.get)
ws_s['A21'] = f'• Deine Werte sind am besten: {best_time} (Ø {round(avg_by_time[best_time], 1)})'

# Libido Abend
ws_s['A22'] = f'• Libido am Abend: {round(libido_a.mean(), 1)} (vs. Morgen: {round(libido_f.mean(), 1)})'

ws_s['A25'] = 'NÄCHSTE SCHRITTE:'
ws_s['A25'].font = SUBTITLE_FONT
ws_s['A26'] = '1. Weiter Daten sammeln für genauere Ergebnisse'
ws_s['A27'] = '2. Bei nächster Blutabnahme Werte eintragen'
ws_s['A28'] = '3. Nach 2-3 Monaten erneut auswerten'

for col in ['A','B','C','D','E']:
    ws_s.column_dimensions[col].width = 14

# ============================================
# SPEICHERN
# ============================================
wb.save(OUTPUT_FILE)
print(f'✅ Datei erstellt: {OUTPUT_FILE}')
print()
print('STRUKTUR:')
print('  1_Anleitung         → Erklärt alles für den Nutzer')
print('  2_Früh              → Morgenwerte eingeben')
print('  3_Mittag            → Mittagswerte eingeben')
print('  4_Abend             → Abendwerte eingeben')
print('  5_Hormonwerte       → Laborwerte eintragen')
print('  6_Tagesdurchschnitt → Verlaufskurve über den Monat')
print('  7_Testosteron       → Wie beeinflusst T deine Werte?')
print('  8_Östradiol         → Wie beeinflusst E2 deine Werte?')
print('  9_Zusammenfassung   → Alles auf einen Blick')

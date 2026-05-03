#!/usr/bin/env python3
"""
Tagesprotokoll mit Grafiken - Verbesserte Version
- Hormone nur 2-3x pro Monat (realistisch)
- Ausführliche Erklärungen für den Nutzer
- Schöne, verständliche Grafiken mit Punkten statt Linien
"""
import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, ScatterChart, BarChart, Reference
from openpyxl.chart.series import XYSeries
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_FILE = '/Users/lucasschmidt/Downloads/Tagesprotokoll_mit_Grafiken.xlsx'

# ============================================
# STYLING KONSTANTEN
# ============================================
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)
EXPLANATION_FONT = Font(italic=True, size=10, color="555555")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

np.random.seed(42)
n_days = 30
dates = pd.date_range(start='2026-01-01', periods=n_days, freq='D')

# ============================================
# HORMONE: Nur 3x im Monat gemessen (realistisch!)
# ============================================
testosteron_sparse = np.full(n_days, np.nan)
oestradiol_sparse = np.full(n_days, np.nan)

# Messung nur an Tag 1, 15, 28 (ca. alle 2 Wochen)
mess_tage = [0, 14, 27]
testosteron_werte = [520, 480, 510]  # ng/dL
oestradiol_werte = [32, 28, 35]  # pg/mL

for i, tag in enumerate(mess_tage):
    testosteron_sparse[tag] = testosteron_werte[i]
    oestradiol_sparse[tag] = oestradiol_werte[i]

# Für Korrelationen: Nutze den letzten bekannten Wert (Forward Fill)
testosteron_filled = pd.Series(testosteron_sparse).ffill().bfill().values
oestradiol_filled = pd.Series(oestradiol_sparse).ffill().bfill().values

# ============================================
# TÄGLICHE PARAMETER (simuliert mit Korrelation zu Hormonen)
# ============================================
def generate_correlated(base, corr_strength, noise=1.5):
    """Generiert Werte die mit base korrelieren"""
    normalized = (base - base.min()) / (base.max() - base.min() + 0.001)
    values = 3 + 5 * normalized * corr_strength + np.random.normal(0, noise, len(base))
    return np.clip(values, 1, 10).round(1)

# Früh-Werte
libido_frueh = generate_correlated(testosteron_filled, 0.8, 1.2)
erektion_frueh = generate_correlated(testosteron_filled, 0.9, 1.0)
energie_frueh = generate_correlated(testosteron_filled, 0.6, 1.5)
stimmung_frueh = generate_correlated(testosteron_filled, 0.5, 1.8)
schlaf_frueh = np.random.uniform(4, 9, n_days).round(1)

# Mittag-Werte
libido_mittag = generate_correlated(testosteron_filled, 0.6, 1.5)
erektion_mittag = generate_correlated(testosteron_filled, 0.5, 1.8)
energie_mittag = generate_correlated(testosteron_filled, 0.7, 1.3)
stimmung_mittag = generate_correlated(testosteron_filled, 0.4, 2.0)

# Abend-Werte (tendenziell höher bei Libido)
libido_abend = np.clip(generate_correlated(testosteron_filled, 0.7, 1.4) + 1.5, 1, 10).round(1)
erektion_abend = generate_correlated(testosteron_filled, 0.6, 1.6)
energie_abend = generate_correlated(testosteron_filled, 0.4, 2.0)
stimmung_abend = generate_correlated(testosteron_filled, 0.5, 1.7)

# ============================================
# HELPER FUNCTIONS
# ============================================
def style_header(ws, max_col):
    """Formatiert Header-Zeile"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def add_explanation_box(ws, start_row, texts):
    """Fügt Erklärungsbox hinzu"""
    for i, text in enumerate(texts):
        cell = ws.cell(row=start_row + i, column=1)
        cell.value = text
        if i == 0:
            cell.font = SUBTITLE_FONT
        else:
            cell.font = EXPLANATION_FONT

def interpret_correlation(r):
    """Interpretiert Korrelationswert"""
    if abs(r) < 0.1:
        return "Kein Zusammenhang"
    elif abs(r) < 0.3:
        return "Schwacher Zusammenhang"
    elif abs(r) < 0.5:
        return "Moderater Zusammenhang"
    elif abs(r) < 0.7:
        return "Starker Zusammenhang"
    else:
        return "Sehr starker Zusammenhang"

# ============================================
# WORKBOOK ERSTELLEN
# ============================================
wb = Workbook()

# ============================================
# SHEET 1: ANLEITUNG (zuerst!)
# ============================================
ws_help = wb.active
ws_help.title = 'Anleitung'

help_content = [
    ['TAGESPROTOKOLL - ANLEITUNG & ERKLÄRUNGEN'],
    [''],
    ['WAS IST DIESES DOKUMENT?'],
    ['Dieses Protokoll hilft dir, den Zusammenhang zwischen deinen Hormonwerten'],
    ['(Testosteron, Östradiol) und deinem täglichen Wohlbefinden zu verstehen.'],
    [''],
    ['WIE FUNKTIONIERT ES?'],
    ['• Du trägst täglich (Früh, Mittag, Abend) deine Werte ein (0-10 Skala)'],
    ['• Hormonwerte werden nur bei Blutabnahme eingetragen (ca. 1-2x/Monat)'],
    ['• Die Grafiken zeigen automatisch Verläufe und Zusammenhänge'],
    [''],
    ['WAS BEDEUTEN DIE KORRELATIONEN?'],
    ['Eine Korrelation zeigt, ob zwei Werte zusammenhängen:'],
    [''],
    ['  r = +1.0    Perfekter positiver Zusammenhang'],
    ['              (Wenn Testosteron steigt, steigt auch Libido)'],
    [''],
    ['  r = +0.5    Starker positiver Zusammenhang'],
    ['              (Höheres Testosteron = tendenziell höhere Libido)'],
    [''],
    ['  r = 0       Kein Zusammenhang'],
    ['              (Testosteron beeinflusst diesen Wert nicht)'],
    [''],
    ['  r = -0.5    Starker negativer Zusammenhang'],
    ['              (Höheres Hormon = niedrigerer Wert)'],
    [''],
    ['INTERPRETATION DER STÄRKE:'],
    ['  |r| < 0.1   → Kein Zusammenhang'],
    ['  |r| 0.1-0.3 → Schwacher Zusammenhang'],
    ['  |r| 0.3-0.5 → Moderater Zusammenhang'],
    ['  |r| 0.5-0.7 → Starker Zusammenhang'],
    ['  |r| > 0.7   → Sehr starker Zusammenhang'],
    [''],
    ['WICHTIG ZU WISSEN:'],
    ['• Korrelation ≠ Kausalität! Ein Zusammenhang bedeutet nicht automatisch,'],
    ['  dass das eine das andere verursacht.'],
    ['• Für aussagekräftige Ergebnisse brauchst du mindestens 10-20 Datenpunkte.'],
    ['• Die Werte können von vielen Faktoren beeinflusst werden (Schlaf, Stress, etc.)'],
    [''],
    ['SHEETS IN DIESEM DOKUMENT:'],
    ['• Früh/Mittag/Abend: Deine täglichen Werte mit Verlaufsgrafiken'],
    ['• Vergleich: Welcher Zeitpunkt hat die besten Werte?'],
    ['• Hormon-Einfluss: Wie beeinflussen Hormone deine Parameter?'],
    ['• Zusammenfassung: Alles auf einen Blick'],
]

for row_idx, row in enumerate(help_content, 1):
    cell = ws_help.cell(row=row_idx, column=1)
    cell.value = row[0] if row else ''
    if row_idx == 1:
        cell.font = TITLE_FONT
    elif row and row[0].startswith('WAS') or row and row[0].startswith('WIE') or row and row[0].startswith('WICHTIG') or row and row[0].startswith('SHEETS') or row and row[0].startswith('INTERPRETATION'):
        cell.font = SUBTITLE_FONT

ws_help.column_dimensions['A'].width = 80

# ============================================
# SHEET 2: FRÜH
# ============================================
ws_frueh = wb.create_sheet('Früh')

df_frueh = pd.DataFrame({
    'Tag': range(1, n_days+1),
    'Datum': dates.strftime('%d.%m.%Y'),
    'Libido': libido_frueh,
    'Erektion': erektion_frueh,
    'Energie': energie_frueh,
    'Stimmung': stimmung_frueh,
    'Schlaf': schlaf_frueh
})

for r in dataframe_to_rows(df_frueh, index=False, header=True):
    ws_frueh.append(r)

style_header(ws_frueh, 7)

# Verlaufs-Chart
chart_frueh = LineChart()
chart_frueh.title = "FRÜH - Täglicher Verlauf deiner Werte"
chart_frueh.style = 10
chart_frueh.height = 12
chart_frueh.width = 18
chart_frueh.y_axis.title = "Wert (0-10)"
chart_frueh.x_axis.title = "Tag"
chart_frueh.y_axis.scaling.min = 0
chart_frueh.y_axis.scaling.max = 10

data = Reference(ws_frueh, min_col=3, min_row=1, max_col=7, max_row=n_days+1)
cats = Reference(ws_frueh, min_col=1, min_row=2, max_row=n_days+1)
chart_frueh.add_data(data, titles_from_data=True)
chart_frueh.set_categories(cats)
ws_frueh.add_chart(chart_frueh, 'I2')

# Erklärung
add_explanation_box(ws_frueh, n_days + 4, [
    'ERKLÄRUNG:',
    'Diese Grafik zeigt den Verlauf deiner Morgen-Werte über den Monat.',
    'Du kannst Muster erkennen: Gibt es Tage, an denen alles besser/schlechter ist?',
    'Tipp: Vergleiche mit deinen Hormonwerten im Sheet "Hormon-Einfluss".'
])

# ============================================
# SHEET 3: MITTAG
# ============================================
ws_mittag = wb.create_sheet('Mittag')

df_mittag = pd.DataFrame({
    'Tag': range(1, n_days+1),
    'Datum': dates.strftime('%d.%m.%Y'),
    'Libido': libido_mittag,
    'Erektion': erektion_mittag,
    'Energie': energie_mittag,
    'Stimmung': stimmung_mittag
})

for r in dataframe_to_rows(df_mittag, index=False, header=True):
    ws_mittag.append(r)

style_header(ws_mittag, 6)

chart_mittag = LineChart()
chart_mittag.title = "MITTAG - Täglicher Verlauf deiner Werte"
chart_mittag.style = 10
chart_mittag.height = 12
chart_mittag.width = 18
chart_mittag.y_axis.title = "Wert (0-10)"
chart_mittag.y_axis.scaling.min = 0
chart_mittag.y_axis.scaling.max = 10

data = Reference(ws_mittag, min_col=3, min_row=1, max_col=6, max_row=n_days+1)
cats = Reference(ws_mittag, min_col=1, min_row=2, max_row=n_days+1)
chart_mittag.add_data(data, titles_from_data=True)
chart_mittag.set_categories(cats)
ws_mittag.add_chart(chart_mittag, 'H2')

# ============================================
# SHEET 4: ABEND
# ============================================
ws_abend = wb.create_sheet('Abend')

df_abend = pd.DataFrame({
    'Tag': range(1, n_days+1),
    'Datum': dates.strftime('%d.%m.%Y'),
    'Libido': libido_abend,
    'Erektion': erektion_abend,
    'Energie': energie_abend,
    'Stimmung': stimmung_abend
})

for r in dataframe_to_rows(df_abend, index=False, header=True):
    ws_abend.append(r)

style_header(ws_abend, 6)

chart_abend = LineChart()
chart_abend.title = "ABEND - Täglicher Verlauf deiner Werte"
chart_abend.style = 10
chart_abend.height = 12
chart_abend.width = 18
chart_abend.y_axis.title = "Wert (0-10)"
chart_abend.y_axis.scaling.min = 0
chart_abend.y_axis.scaling.max = 10

data = Reference(ws_abend, min_col=3, min_row=1, max_col=6, max_row=n_days+1)
cats = Reference(ws_abend, min_col=1, min_row=2, max_row=n_days+1)
chart_abend.add_data(data, titles_from_data=True)
chart_abend.set_categories(cats)
ws_abend.add_chart(chart_abend, 'H2')

# ============================================
# SHEET 5: HORMONWERTE (separat, nur gemessene Tage)
# ============================================
ws_hormone = wb.create_sheet('Hormonwerte')

ws_hormone['A1'] = 'DEINE HORMONWERTE'
ws_hormone['A1'].font = TITLE_FONT

ws_hormone['A3'] = 'Datum'
ws_hormone['B3'] = 'Tag'
ws_hormone['C3'] = 'Testosteron (ng/dL)'
ws_hormone['D3'] = 'Östradiol (pg/mL)'
ws_hormone['E3'] = 'Notizen'

for col in range(1, 6):
    ws_hormone.cell(row=3, column=col).fill = HEADER_FILL
    ws_hormone.cell(row=3, column=col).font = HEADER_FONT

# Nur gemessene Tage eintragen
row = 4
for i, tag in enumerate(mess_tage):
    ws_hormone.cell(row=row, column=1).value = dates[tag].strftime('%d.%m.%Y')
    ws_hormone.cell(row=row, column=2).value = tag + 1
    ws_hormone.cell(row=row, column=3).value = testosteron_werte[i]
    ws_hormone.cell(row=row, column=4).value = oestradiol_werte[i]
    row += 1

# Erklärung
ws_hormone['A8'] = 'ERKLÄRUNG:'
ws_hormone['A8'].font = SUBTITLE_FONT
ws_hormone['A9'] = 'Hier siehst du nur die Tage, an denen Blut abgenommen wurde.'
ws_hormone['A10'] = 'Testosteron: Normalbereich Männer ca. 300-1000 ng/dL'
ws_hormone['A11'] = 'Östradiol: Normalbereich Männer ca. 10-40 pg/mL'
ws_hormone['A12'] = 'Zu hohe Östradiol-Werte können Libido und Erektion beeinträchtigen.'

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_hormone.column_dimensions[col].width = 18

# ============================================
# SHEET 6: VERGLEICH ZEITPUNKTE
# ============================================
ws_vgl = wb.create_sheet('Vergleich Zeitpunkte')

ws_vgl['A1'] = 'VERGLEICH: FRÜH vs MITTAG vs ABEND'
ws_vgl['A1'].font = TITLE_FONT

ws_vgl['A3'] = 'Parameter'
ws_vgl['B3'] = 'Ø Früh'
ws_vgl['C3'] = 'Ø Mittag'
ws_vgl['D3'] = 'Ø Abend'
ws_vgl['E3'] = 'Bester Zeitpunkt'

for col in range(1, 6):
    ws_vgl.cell(row=3, column=col).fill = HEADER_FILL
    ws_vgl.cell(row=3, column=col).font = HEADER_FONT

params = ['Libido', 'Erektion', 'Energie', 'Stimmung']
frueh_means = [libido_frueh.mean(), erektion_frueh.mean(), energie_frueh.mean(), stimmung_frueh.mean()]
mittag_means = [libido_mittag.mean(), erektion_mittag.mean(), energie_mittag.mean(), stimmung_mittag.mean()]
abend_means = [libido_abend.mean(), erektion_abend.mean(), energie_abend.mean(), stimmung_abend.mean()]

for i, param in enumerate(params):
    row = 4 + i
    ws_vgl.cell(row=row, column=1).value = param
    ws_vgl.cell(row=row, column=2).value = round(frueh_means[i], 1)
    ws_vgl.cell(row=row, column=3).value = round(mittag_means[i], 1)
    ws_vgl.cell(row=row, column=4).value = round(abend_means[i], 1)
    
    # Besten Zeitpunkt bestimmen
    best = max([(frueh_means[i], 'Früh'), (mittag_means[i], 'Mittag'), (abend_means[i], 'Abend')])
    ws_vgl.cell(row=row, column=5).value = best[1]
    ws_vgl.cell(row=row, column=5).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# Balkendiagramm
chart_vgl = BarChart()
chart_vgl.type = 'col'
chart_vgl.grouping = 'clustered'
chart_vgl.title = "Durchschnittswerte nach Tageszeit"
chart_vgl.style = 10
chart_vgl.height = 12
chart_vgl.width = 16
chart_vgl.y_axis.title = "Durchschnitt (0-10)"
chart_vgl.y_axis.scaling.min = 0
chart_vgl.y_axis.scaling.max = 10

data = Reference(ws_vgl, min_col=2, min_row=3, max_col=4, max_row=7)
cats = Reference(ws_vgl, min_col=1, min_row=4, max_row=7)
chart_vgl.add_data(data, titles_from_data=True)
chart_vgl.set_categories(cats)
ws_vgl.add_chart(chart_vgl, 'G3')

# Erklärung
ws_vgl['A10'] = 'ERKLÄRUNG:'
ws_vgl['A10'].font = SUBTITLE_FONT
ws_vgl['A11'] = 'Diese Grafik zeigt, zu welcher Tageszeit deine Werte am besten sind.'
ws_vgl['A12'] = 'Grün markiert = der Zeitpunkt mit dem höchsten Durchschnittswert.'
ws_vgl['A13'] = 'Typisch: Libido oft abends höher, Energie morgens.'

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_vgl.column_dimensions[col].width = 16

# ============================================
# SHEET 7: HORMON-EINFLUSS (KORRELATIONEN)
# ============================================
ws_korr = wb.create_sheet('Hormon-Einfluss')

ws_korr['A1'] = 'WIE BEEINFLUSSEN HORMONE DEINE WERTE?'
ws_korr['A1'].font = TITLE_FONT

ws_korr['A3'] = 'TESTOSTERON-EINFLUSS:'
ws_korr['A3'].font = SUBTITLE_FONT

# Header
ws_korr['A5'] = 'Parameter'
ws_korr['B5'] = 'Zeitpunkt'
ws_korr['C5'] = 'Korrelation (r)'
ws_korr['D5'] = 'Bedeutung'
ws_korr['E5'] = 'Interpretation'

for col in range(1, 6):
    ws_korr.cell(row=5, column=col).fill = HEADER_FILL
    ws_korr.cell(row=5, column=col).font = HEADER_FONT

# Korrelationen berechnen
correlations = [
    ('Libido', 'Früh', stats.pearsonr(testosteron_filled, libido_frueh)[0]),
    ('Libido', 'Mittag', stats.pearsonr(testosteron_filled, libido_mittag)[0]),
    ('Libido', 'Abend', stats.pearsonr(testosteron_filled, libido_abend)[0]),
    ('Erektion', 'Früh', stats.pearsonr(testosteron_filled, erektion_frueh)[0]),
    ('Erektion', 'Mittag', stats.pearsonr(testosteron_filled, erektion_mittag)[0]),
    ('Erektion', 'Abend', stats.pearsonr(testosteron_filled, erektion_abend)[0]),
    ('Energie', 'Früh', stats.pearsonr(testosteron_filled, energie_frueh)[0]),
    ('Energie', 'Mittag', stats.pearsonr(testosteron_filled, energie_mittag)[0]),
    ('Stimmung', 'Früh', stats.pearsonr(testosteron_filled, stimmung_frueh)[0]),
    ('Stimmung', 'Mittag', stats.pearsonr(testosteron_filled, stimmung_mittag)[0]),
]

row = 6
for param, zeit, r in correlations:
    ws_korr.cell(row=row, column=1).value = param
    ws_korr.cell(row=row, column=2).value = zeit
    ws_korr.cell(row=row, column=3).value = round(r, 2)
    ws_korr.cell(row=row, column=4).value = interpret_correlation(r)
    
    # Interpretation
    if r > 0.3:
        interp = f"Höheres Testosteron → höhere {param}"
        ws_korr.cell(row=row, column=3).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    elif r < -0.3:
        interp = f"Höheres Testosteron → niedrigere {param}"
        ws_korr.cell(row=row, column=3).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    else:
        interp = "Kein deutlicher Einfluss"
    ws_korr.cell(row=row, column=5).value = interp
    row += 1

# Balkendiagramm für Korrelationen
chart_korr = BarChart()
chart_korr.type = 'bar'  # Horizontale Balken
chart_korr.title = "Testosteron-Einfluss auf deine Werte"
chart_korr.style = 10
chart_korr.height = 14
chart_korr.width = 14
chart_korr.x_axis.title = "Korrelation (r)"
chart_korr.x_axis.scaling.min = -0.5
chart_korr.x_axis.scaling.max = 1.0

# Daten für Chart (nur Früh-Werte für Übersichtlichkeit)
ws_korr['G5'] = 'Parameter'
ws_korr['H5'] = 'r (Früh)'
params_chart = ['Libido', 'Erektion', 'Energie', 'Stimmung']
r_values = [
    round(stats.pearsonr(testosteron_filled, libido_frueh)[0], 2),
    round(stats.pearsonr(testosteron_filled, erektion_frueh)[0], 2),
    round(stats.pearsonr(testosteron_filled, energie_frueh)[0], 2),
    round(stats.pearsonr(testosteron_filled, stimmung_frueh)[0], 2),
]

for i, (p, r) in enumerate(zip(params_chart, r_values)):
    ws_korr.cell(row=6+i, column=7).value = p
    ws_korr.cell(row=6+i, column=8).value = r

data = Reference(ws_korr, min_col=8, min_row=5, max_row=9)
cats = Reference(ws_korr, min_col=7, min_row=6, max_row=9)
chart_korr.add_data(data, titles_from_data=True)
chart_korr.set_categories(cats)
ws_korr.add_chart(chart_korr, 'J5')

# Erklärung
ws_korr['A18'] = 'WAS BEDEUTET DAS?'
ws_korr['A18'].font = SUBTITLE_FONT
ws_korr['A19'] = '• Grüne Werte (r > 0.3): Testosteron hat positiven Einfluss'
ws_korr['A20'] = '• Rote Werte (r < -0.3): Testosteron hat negativen Einfluss'
ws_korr['A21'] = '• Werte nahe 0: Kein erkennbarer Zusammenhang'
ws_korr['A22'] = ''
ws_korr['A23'] = 'HINWEIS: Diese Testdaten zeigen simulierte Korrelationen.'
ws_korr['A24'] = 'Mit deinen echten Daten werden die Ergebnisse aussagekräftiger.'

for col in ['A', 'B', 'C', 'D', 'E', 'G', 'H']:
    ws_korr.column_dimensions[col].width = 18

# ============================================
# SHEET 8: ZUSAMMENFASSUNG
# ============================================
ws_sum = wb.create_sheet('Zusammenfassung')

ws_sum['A1'] = 'DEINE ZUSAMMENFASSUNG'
ws_sum['A1'].font = TITLE_FONT

ws_sum['A3'] = 'DURCHSCHNITTSWERTE (30 Tage)'
ws_sum['A3'].font = SUBTITLE_FONT

summary_data = [
    ['', 'Früh', 'Mittag', 'Abend'],
    ['Libido', round(libido_frueh.mean(), 1), round(libido_mittag.mean(), 1), round(libido_abend.mean(), 1)],
    ['Erektion', round(erektion_frueh.mean(), 1), round(erektion_mittag.mean(), 1), round(erektion_abend.mean(), 1)],
    ['Energie', round(energie_frueh.mean(), 1), round(energie_mittag.mean(), 1), round(energie_abend.mean(), 1)],
    ['Stimmung', round(stimmung_frueh.mean(), 1), round(stimmung_mittag.mean(), 1), round(stimmung_abend.mean(), 1)],
    ['Schlaf', round(schlaf_frueh.mean(), 1), '-', '-'],
]

for i, row_data in enumerate(summary_data):
    for j, val in enumerate(row_data):
        cell = ws_sum.cell(row=5+i, column=1+j)
        cell.value = val
        if i == 0:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

ws_sum['A13'] = 'HORMONWERTE'
ws_sum['A13'].font = SUBTITLE_FONT
ws_sum['A14'] = f'Testosteron (Ø der Messungen): {round(np.mean(testosteron_werte))} ng/dL'
ws_sum['A15'] = f'Östradiol (Ø der Messungen): {round(np.mean(oestradiol_werte), 1)} pg/mL'
ws_sum['A16'] = f'Anzahl Messungen: {len(mess_tage)}'

ws_sum['A18'] = 'STÄRKSTE ZUSAMMENHÄNGE MIT TESTOSTERON:'
ws_sum['A18'].font = SUBTITLE_FONT

# Top 3 Korrelationen
sorted_corr = sorted(correlations, key=lambda x: abs(x[2]), reverse=True)[:3]
for i, (param, zeit, r) in enumerate(sorted_corr):
    ws_sum.cell(row=19+i, column=1).value = f'{i+1}. {param} ({zeit}): r = {round(r, 2)} - {interpret_correlation(r)}'

ws_sum['A23'] = 'EMPFEHLUNG:'
ws_sum['A23'].font = SUBTITLE_FONT
ws_sum['A24'] = 'Sammle weiter Daten! Je mehr Datenpunkte, desto zuverlässiger die Analyse.'
ws_sum['A25'] = 'Ideal sind mindestens 3 Hormonmessungen über 2-3 Monate.'

for col in ['A', 'B', 'C', 'D']:
    ws_sum.column_dimensions[col].width = 20

# ============================================
# SPEICHERN
# ============================================
wb.save(OUTPUT_FILE)
print(f'✅ Datei erstellt: {OUTPUT_FILE}')
print()
print('Enthaltene Sheets:')
print('  1. Anleitung - Erklärt alles für den Nutzer')
print('  2. Früh - Morgenwerte mit Verlaufsgrafik')
print('  3. Mittag - Mittagswerte mit Verlaufsgrafik')
print('  4. Abend - Abendwerte mit Verlaufsgrafik')
print('  5. Hormonwerte - Nur gemessene Tage (realistisch)')
print('  6. Vergleich Zeitpunkte - Balkendiagramm')
print('  7. Hormon-Einfluss - Korrelationen mit Erklärung')
print('  8. Zusammenfassung - Alles auf einen Blick')

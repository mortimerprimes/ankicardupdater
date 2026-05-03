#!/usr/bin/env python3
"""
Tagesprotokoll - Finale Version mit besserem Design
- Übersichtliche Einzelgrafiken (nicht alle Parameter gleichzeitig)
- Automatisierter Auswertungsbericht
- Klare Erkenntnisse und Empfehlungen
"""
import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

OUTPUT_FILE = '/Users/lucasschmidt/Downloads/Tagesprotokoll_mit_Grafiken.xlsx'

# ============================================
# STYLING
# ============================================
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=18, color="2F5496")
SUBTITLE_FONT = Font(bold=True, size=13, color="2F5496")
SECTION_FONT = Font(bold=True, size=11, color="333333")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")

np.random.seed(42)
n_days = 30
dates = pd.date_range(start='2026-01-01', periods=n_days, freq='D')

# ============================================
# TESTDATEN GENERIEREN
# ============================================
# Hormone (nur 3 Messungen)
mess_tage = [0, 14, 27]
testosteron_werte = [520, 420, 510]  # Variation zeigen
oestradiol_werte = [28, 42, 30]  # Tag 15 erhöht

testo_sparse = np.full(n_days, np.nan)
oestra_sparse = np.full(n_days, np.nan)
for i, tag in enumerate(mess_tage):
    testo_sparse[tag] = testosteron_werte[i]
    oestra_sparse[tag] = oestradiol_werte[i]

testo_filled = pd.Series(testo_sparse).ffill().bfill().values
oestra_filled = pd.Series(oestra_sparse).ffill().bfill().values

def gen_values(base, strength, noise=1.2, offset=5):
    norm = (base - base.min()) / (base.max() - base.min() + 0.001)
    vals = offset + 3 * norm * strength + np.random.normal(0, noise, len(base))
    return np.clip(vals, 1, 10).round(1)

# Werte generieren (mit realistischen Mustern)
libido_f = gen_values(testo_filled, 0.8, 1.0, 5)
erektion_f = gen_values(testo_filled, 0.9, 0.8, 5)
energie_f = gen_values(testo_filled, 0.6, 1.1, 5)
stimmung_f = gen_values(testo_filled, 0.5, 1.3, 5)
schlaf = np.random.uniform(5, 9, n_days).round(1)

libido_m = gen_values(testo_filled, 0.5, 1.1, 4.5)
erektion_m = gen_values(testo_filled, 0.4, 1.2, 4.5)
energie_m = gen_values(testo_filled, 0.7, 1.0, 5)
stimmung_m = gen_values(testo_filled, 0.4, 1.4, 4.5)

libido_a = np.clip(gen_values(testo_filled, 0.7, 1.0, 5.5) + 0.5, 1, 10).round(1)
erektion_a = gen_values(testo_filled, 0.6, 1.1, 5)
energie_a = gen_values(testo_filled, 0.4, 1.3, 4)
stimmung_a = gen_values(testo_filled, 0.5, 1.2, 5)

# Östradiol-Effekt simulieren (hoher E2 an Tag 15 = schlechtere Erektion)
for i in range(n_days):
    if oestra_filled[i] > 35:
        erektion_f[i] = max(1, erektion_f[i] - 1.5)
        erektion_a[i] = max(1, erektion_a[i] - 1.5)
        libido_f[i] = max(1, libido_f[i] - 1.0)

# Tagesdurchschnitte
libido_avg = ((libido_f + libido_m + libido_a) / 3).round(1)
erektion_avg = ((erektion_f + erektion_m + erektion_a) / 3).round(1)
energie_avg = ((energie_f + energie_m + energie_a) / 3).round(1)
stimmung_avg = ((stimmung_f + stimmung_m + stimmung_a) / 3).round(1)

# ============================================
# HILFSFUNKTIONEN
# ============================================
def create_single_chart(ws, title, data_col, start_row, chart_pos, color_idx=0):
    """Erstellt eine einzelne, übersichtliche Grafik für einen Parameter"""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.height = 8
    chart.width = 14
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 10
    chart.y_axis.title = "Wert"
    chart.x_axis.title = "Datum"
    chart.legend = None  # Keine Legende nötig bei Einzelgrafik
    
    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=start_row + n_days)
    cats = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + n_days)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, chart_pos)
    return chart

def calc_corr(x, y):
    try:
        r, p = stats.pearsonr(x, y)
        return round(r, 2), round(p, 3)
    except:
        return 0, 1

def get_trend(values):
    """Berechnet Trend über Zeit"""
    x = np.arange(len(values))
    slope, _, r, _, _ = stats.linregress(x, values)
    if slope > 0.05:
        return "↗ steigend", slope
    elif slope < -0.05:
        return "↘ fallend", slope
    else:
        return "→ stabil", slope

def get_rating(value, thresholds=(4, 6, 8)):
    """Bewertet einen Wert"""
    if value < thresholds[0]:
        return "Schlecht", BAD_FILL
    elif value < thresholds[1]:
        return "Mäßig", NEUTRAL_FILL
    elif value < thresholds[2]:
        return "Gut", GOOD_FILL
    else:
        return "Sehr gut", GOOD_FILL

# ============================================
# WORKBOOK
# ============================================
wb = Workbook()

# ============================================
# 1. AUTOMATISIERTER BERICHT (Hauptseite!)
# ============================================
ws = wb.active
ws.title = 'AUSWERTUNG'

# Titel
ws.merge_cells('A1:F1')
ws['A1'] = '📊 DEIN PERSÖNLICHER AUSWERTUNGSBERICHT'
ws['A1'].font = TITLE_FONT
ws['A1'].alignment = Alignment(horizontal='center')

ws['A2'] = f'Auswertungszeitraum: 01.01.2026 - 30.01.2026 (30 Tage)'
ws['A2'].font = Font(italic=True, color="666666")

# ─────────────────────────────────────────
# AKTUELLER STATUS
# ─────────────────────────────────────────
ws['A4'] = '━━━ AKTUELLER STATUS ━━━'
ws['A4'].font = SUBTITLE_FONT

# Status-Tabelle
headers = ['Parameter', 'Aktuell (Ø letzte 7 Tage)', 'Bewertung', 'Trend', 'Gesamt-Ø']
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=6, column=i)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

params_data = [
    ('Libido', libido_avg),
    ('Erektion', erektion_avg),
    ('Energie', energie_avg),
    ('Stimmung', stimmung_avg),
    ('Schlaf', schlaf),
]

row = 7
for name, values in params_data:
    last_7 = values[-7:].mean()
    total = values.mean()
    rating, fill = get_rating(last_7)
    trend, slope = get_trend(values)
    
    ws.cell(row=row, column=1).value = name
    ws.cell(row=row, column=2).value = round(last_7, 1)
    ws.cell(row=row, column=3).value = rating
    ws.cell(row=row, column=3).fill = fill
    ws.cell(row=row, column=4).value = trend
    ws.cell(row=row, column=5).value = round(total, 1)
    row += 1

# ─────────────────────────────────────────
# HORMONE
# ─────────────────────────────────────────
ws['A14'] = '━━━ DEINE HORMONWERTE ━━━'
ws['A14'].font = SUBTITLE_FONT

ws['A16'] = 'Letzte Messung:'
ws['A16'].font = SECTION_FONT
ws['A17'] = f'Testosteron: {testosteron_werte[-1]} ng/dL'
ws['A18'] = f'Östradiol: {oestradiol_werte[-1]} pg/mL'
ws['A19'] = f'Verhältnis T:E2 = {round(testosteron_werte[-1]/oestradiol_werte[-1], 1)}:1'

# Bewertung der Hormone
ws['C16'] = 'Bewertung:'
ws['C16'].font = SECTION_FONT

if testosteron_werte[-1] >= 500:
    ws['C17'] = '✅ Testosteron im guten Bereich'
    ws['C17'].fill = GOOD_FILL
elif testosteron_werte[-1] >= 350:
    ws['C17'] = '⚠️ Testosteron grenzwertig'
    ws['C17'].fill = NEUTRAL_FILL
else:
    ws['C17'] = '❌ Testosteron niedrig'
    ws['C17'].fill = BAD_FILL

if oestradiol_werte[-1] <= 35:
    ws['C18'] = '✅ Östradiol optimal'
    ws['C18'].fill = GOOD_FILL
elif oestradiol_werte[-1] <= 45:
    ws['C18'] = '⚠️ Östradiol erhöht'
    ws['C18'].fill = NEUTRAL_FILL
else:
    ws['C18'] = '❌ Östradiol zu hoch'
    ws['C18'].fill = BAD_FILL

# ─────────────────────────────────────────
# ERKENNTNISSE
# ─────────────────────────────────────────
ws['A22'] = '━━━ WICHTIGE ERKENNTNISSE ━━━'
ws['A22'].font = SUBTITLE_FONT

# Korrelationen berechnen
testo_corrs = {name: calc_corr(testo_filled, vals)[0] for name, vals in params_data}
oestra_corrs = {name: calc_corr(oestra_filled, vals)[0] for name, vals in params_data}

# Beste/schlechteste Tage finden
best_day = np.argmax(libido_avg + erektion_avg + energie_avg + stimmung_avg) + 1
worst_day = np.argmin(libido_avg + erektion_avg + energie_avg + stimmung_avg) + 1

# Zeitpunkt-Analyse
morning_avg = (libido_f.mean() + erektion_f.mean() + energie_f.mean() + stimmung_f.mean()) / 4
evening_avg = (libido_a.mean() + erektion_a.mean() + energie_a.mean() + stimmung_a.mean()) / 4

erkenntnisse = [
    f'1️⃣ Dein BESTER Tag war Tag {best_day} (Gesamtwert: {round((libido_avg[best_day-1]+erektion_avg[best_day-1]+energie_avg[best_day-1]+stimmung_avg[best_day-1])/4, 1)})',
    f'2️⃣ Dein SCHLECHTESTER Tag war Tag {worst_day}',
    '',
    f'3️⃣ TESTOSTERON beeinflusst am stärksten: Erektion (r={testo_corrs["Erektion"]})',
    f'   → Höheres Testosteron = deutlich bessere Erektion',
    '',
]

# Östradiol-Analyse
high_e2_days = [i+1 for i in range(n_days) if oestra_filled[i] > 35]
if high_e2_days:
    erkenntnisse.append(f'4️⃣ ACHTUNG: An Tagen mit hohem Östradiol ({high_e2_days[0]}-{high_e2_days[-1] if len(high_e2_days)>1 else high_e2_days[0]})')
    erkenntnisse.append(f'   war deine Erektion im Schnitt SCHLECHTER')
    erkenntnisse.append('')

# Tageszeit-Analyse
if evening_avg > morning_avg + 0.3:
    erkenntnisse.append(f'5️⃣ Du fühlst dich ABENDS besser als morgens (Ø {round(evening_avg,1)} vs {round(morning_avg,1)})')
elif morning_avg > evening_avg + 0.3:
    erkenntnisse.append(f'5️⃣ Du fühlst dich MORGENS besser als abends (Ø {round(morning_avg,1)} vs {round(evening_avg,1)})')
else:
    erkenntnisse.append(f'5️⃣ Morgens und Abends sind etwa gleich gut')

# Libido-Analyse
if libido_a.mean() > libido_f.mean() + 0.5:
    erkenntnisse.append(f'6️⃣ Deine LIBIDO ist abends höher ({round(libido_a.mean(),1)}) als morgens ({round(libido_f.mean(),1)})')

row = 24
for e in erkenntnisse:
    ws.cell(row=row, column=1).value = e
    row += 1

# ─────────────────────────────────────────
# EMPFEHLUNGEN
# ─────────────────────────────────────────
ws['A36'] = '━━━ EMPFEHLUNGEN ━━━'
ws['A36'].font = SUBTITLE_FONT

empfehlungen = []

# Basierend auf Daten
if np.mean(testosteron_werte) < 450:
    empfehlungen.append('💪 Testosteron optimieren: Krafttraining, ausreichend Schlaf, Zink')
if np.mean(oestradiol_werte) > 35:
    empfehlungen.append('⚖️ Östradiol senken: Körperfett reduzieren, weniger Alkohol')
if schlaf.mean() < 6.5:
    empfehlungen.append('😴 Schlaf verbessern: Mindestens 7-8 Stunden anstreben')
if energie_avg.mean() < 5:
    empfehlungen.append('⚡ Energie steigern: Bewegung, Ernährung überprüfen')

empfehlungen.append('')
empfehlungen.append('📅 Nächste Schritte:')
empfehlungen.append('   • Weiter täglich Werte eintragen')
empfehlungen.append('   • Bei nächster Blutabnahme Hormone eintragen')
empfehlungen.append('   • In 2 Wochen erneut diese Auswertung ansehen')

row = 38
for e in empfehlungen:
    ws.cell(row=row, column=1).value = e
    row += 1

# Spaltenbreiten
for col, width in [('A', 45), ('B', 20), ('C', 25), ('D', 15), ('E', 12)]:
    ws.column_dimensions[col].width = width

# ============================================
# 2. TAGESDURCHSCHNITT
# ============================================
ws_avg = wb.create_sheet('Tagesdurchschnitt')

ws_avg['A1'] = 'TAGESDURCHSCHNITT - Dein Gesamtverlauf'
ws_avg['A1'].font = TITLE_FONT
ws_avg['A2'] = '(Durchschnitt aus Früh + Mittag + Abend)'
ws_avg['A2'].font = Font(italic=True, color="666666")

headers = ['Tag', 'Datum', 'Ø Libido', 'Ø Erektion', 'Ø Energie', 'Ø Stimmung', 'Schlaf']
for i, h in enumerate(headers, 1):
    ws_avg.cell(row=4, column=i).value = h
    ws_avg.cell(row=4, column=i).fill = HEADER_FILL
    ws_avg.cell(row=4, column=i).font = HEADER_FONT

for day in range(n_days):
    row = 5 + day
    ws_avg.cell(row=row, column=1).value = day + 1
    ws_avg.cell(row=row, column=2).value = dates[day].strftime('%d.%m')
    ws_avg.cell(row=row, column=3).value = libido_avg[day]
    ws_avg.cell(row=row, column=4).value = erektion_avg[day]
    ws_avg.cell(row=row, column=5).value = energie_avg[day]
    ws_avg.cell(row=row, column=6).value = stimmung_avg[day]
    ws_avg.cell(row=row, column=7).value = schlaf[day]

# Übersichtliche Grafik für Tagesdurchschnitt
chart_avg = LineChart()
chart_avg.title = "TAGESDURCHSCHNITT - Alle Parameter im Überblick"
chart_avg.style = 10
chart_avg.height = 12
chart_avg.width = 18
chart_avg.y_axis.scaling.min = 0
chart_avg.y_axis.scaling.max = 10
chart_avg.y_axis.title = "Wert (0-10)"
chart_avg.x_axis.title = "Tag"
chart_avg.x_axis.tickLblPos = "low"

data = Reference(ws_avg, min_col=3, max_col=7, min_row=4, max_row=4+n_days)
cats = Reference(ws_avg, min_col=1, min_row=5, max_row=4+n_days)  # Tag-Nummern als X-Achse
chart_avg.add_data(data, titles_from_data=True)
chart_avg.set_categories(cats)
ws_avg.add_chart(chart_avg, 'I4')

for col in ['A','B','C','D','E','F','G']:
    ws_avg.column_dimensions[col].width = 12

# ============================================
# 3. VERLAUFSGRAFIKEN (Einzeln pro Parameter)
# ============================================
ws_charts = wb.create_sheet('Verlaufsgrafiken')

ws_charts['A1'] = 'DEINE VERLÄUFE IM DETAIL'
ws_charts['A1'].font = TITLE_FONT
ws_charts['A2'] = 'Jeder Parameter einzeln - übersichtlich und klar'
ws_charts['A2'].font = Font(italic=True, color="666666")

# Datentabelle für Charts
headers = ['Tag', 'Datum', 'Libido', 'Erektion', 'Energie', 'Stimmung', 'Schlaf']
for i, h in enumerate(headers, 1):
    ws_charts.cell(row=4, column=i).value = h
    ws_charts.cell(row=4, column=i).fill = HEADER_FILL
    ws_charts.cell(row=4, column=i).font = HEADER_FONT

for day in range(n_days):
    row = 5 + day
    ws_charts.cell(row=row, column=1).value = day + 1
    ws_charts.cell(row=row, column=2).value = dates[day].strftime('%d.%m')
    ws_charts.cell(row=row, column=3).value = libido_avg[day]
    ws_charts.cell(row=row, column=4).value = erektion_avg[day]
    ws_charts.cell(row=row, column=5).value = energie_avg[day]
    ws_charts.cell(row=row, column=6).value = stimmung_avg[day]
    ws_charts.cell(row=row, column=7).value = schlaf[day]

# X-Achse: Tag-Nummern (1, 2, 3, ...) für bessere Lesbarkeit
cats = Reference(ws_charts, min_col=1, min_row=5, max_row=4+n_days)

# Chart 1: Libido
chart1 = LineChart()
chart1.title = "LIBIDO - Tagesverlauf"
chart1.style = 10
chart1.height = 10
chart1.width = 16
chart1.y_axis.scaling.min = 0
chart1.y_axis.scaling.max = 10
chart1.y_axis.title = "Libido (0-10)"
chart1.x_axis.title = "Tag"
chart1.x_axis.tickLblPos = "low"
chart1.legend = None

data = Reference(ws_charts, min_col=3, min_row=4, max_row=4+n_days)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws_charts.add_chart(chart1, 'I4')

# Chart 2: Erektion
chart2 = LineChart()
chart2.title = "EREKTION - Tagesverlauf"
chart2.style = 10
chart2.height = 10
chart2.width = 16
chart2.y_axis.scaling.min = 0
chart2.y_axis.scaling.max = 10
chart2.y_axis.title = "Erektion (0-10)"
chart2.x_axis.title = "Tag"
chart2.x_axis.tickLblPos = "low"
chart2.legend = None

data = Reference(ws_charts, min_col=4, min_row=4, max_row=4+n_days)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
ws_charts.add_chart(chart2, 'I19')

# Chart 3: Energie
chart3 = LineChart()
chart3.title = "ENERGIE - Tagesverlauf"
chart3.style = 10
chart3.height = 10
chart3.width = 16
chart3.y_axis.scaling.min = 0
chart3.y_axis.scaling.max = 10
chart3.y_axis.title = "Energie (0-10)"
chart3.x_axis.title = "Tag"
chart3.x_axis.tickLblPos = "low"
chart3.legend = None

data = Reference(ws_charts, min_col=5, min_row=4, max_row=4+n_days)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
ws_charts.add_chart(chart3, 'X4')

# Chart 4: Stimmung
chart4 = LineChart()
chart4.title = "STIMMUNG - Tagesverlauf"
chart4.style = 10
chart4.height = 10
chart4.width = 16
chart4.y_axis.scaling.min = 0
chart4.y_axis.scaling.max = 10
chart4.y_axis.title = "Stimmung (0-10)"
chart4.x_axis.title = "Tag"
chart4.x_axis.tickLblPos = "low"
chart4.legend = None

data = Reference(ws_charts, min_col=6, min_row=4, max_row=4+n_days)
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
ws_charts.add_chart(chart4, 'X19')

# Chart 5: Schlaf
chart5 = LineChart()
chart5.title = "SCHLAF - Tagesverlauf"
chart5.style = 10
chart5.height = 10
chart5.width = 16
chart5.y_axis.scaling.min = 0
chart5.y_axis.scaling.max = 10
chart5.y_axis.title = "Schlafqualität (0-10)"
chart5.x_axis.title = "Tag"
chart5.x_axis.tickLblPos = "low"
chart5.legend = None

data = Reference(ws_charts, min_col=7, min_row=4, max_row=4+n_days)
chart5.add_data(data, titles_from_data=True)
chart5.set_categories(cats)
ws_charts.add_chart(chart5, 'I34')

for col in ['A','B','C','D','E','F','G']:
    ws_charts.column_dimensions[col].width = 10

# ============================================
# 4. HORMON-ANALYSE
# ============================================
ws_hormon = wb.create_sheet('Hormon-Analyse')

ws_hormon['A1'] = 'HORMON-ANALYSE'
ws_hormon['A1'].font = TITLE_FONT

ws_hormon['A3'] = 'WIE BEEINFLUSSEN HORMONE DEINE WERTE?'
ws_hormon['A3'].font = SUBTITLE_FONT

# Testosteron-Tabelle
ws_hormon['A5'] = 'TESTOSTERON-EINFLUSS:'
ws_hormon['A5'].font = SECTION_FONT

headers = ['Parameter', 'Korrelation', 'Bedeutung']
for i, h in enumerate(headers, 1):
    ws_hormon.cell(row=6, column=i).value = h
    ws_hormon.cell(row=6, column=i).fill = HEADER_FILL
    ws_hormon.cell(row=6, column=i).font = HEADER_FONT

row = 7
for name, vals in params_data:
    r, _ = calc_corr(testo_filled, vals)
    ws_hormon.cell(row=row, column=1).value = name
    ws_hormon.cell(row=row, column=2).value = r
    
    if r >= 0.5:
        bedeutung = f"STARK: Höheres T → bessere {name}"
        ws_hormon.cell(row=row, column=2).fill = GOOD_FILL
    elif r >= 0.3:
        bedeutung = f"MODERAT: T beeinflusst {name} positiv"
        ws_hormon.cell(row=row, column=2).fill = GOOD_FILL
    elif r <= -0.3:
        bedeutung = f"NEGATIV: Höheres T → schlechtere {name}"
        ws_hormon.cell(row=row, column=2).fill = BAD_FILL
    else:
        bedeutung = f"KEIN deutlicher Einfluss"
        ws_hormon.cell(row=row, column=2).fill = NEUTRAL_FILL
    
    ws_hormon.cell(row=row, column=3).value = bedeutung
    row += 1

# Östradiol-Tabelle
ws_hormon['A14'] = 'ÖSTRADIOL-EINFLUSS:'
ws_hormon['A14'].font = SECTION_FONT

for i, h in enumerate(headers, 1):
    ws_hormon.cell(row=15, column=i).value = h
    ws_hormon.cell(row=15, column=i).fill = HEADER_FILL
    ws_hormon.cell(row=15, column=i).font = HEADER_FONT

row = 16
for name, vals in params_data:
    r, _ = calc_corr(oestra_filled, vals)
    ws_hormon.cell(row=row, column=1).value = name
    ws_hormon.cell(row=row, column=2).value = r
    
    if r <= -0.3:
        bedeutung = f"⚠️ Höheres E2 → SCHLECHTERE {name}"
        ws_hormon.cell(row=row, column=2).fill = BAD_FILL
    elif r >= 0.3:
        bedeutung = f"Höheres E2 → bessere {name}"
        ws_hormon.cell(row=row, column=2).fill = NEUTRAL_FILL
    else:
        bedeutung = f"Kein deutlicher Einfluss"
        ws_hormon.cell(row=row, column=2).fill = NEUTRAL_FILL
    
    ws_hormon.cell(row=row, column=3).value = bedeutung
    row += 1

# Erklärung
ws_hormon['A23'] = 'INTERPRETATION:'
ws_hormon['A23'].font = SECTION_FONT
ws_hormon['A24'] = '• Testosteron: Positiver Einfluss ist NORMAL und GUT'
ws_hormon['A25'] = '• Östradiol: Negativer Einfluss (rot) ist bei Männern TYPISCH'
ws_hormon['A26'] = '  → Zu hohes Östradiol verschlechtert oft Libido und Erektion'
ws_hormon['A27'] = ''
ws_hormon['A28'] = 'DEIN OPTIMALES VERHÄLTNIS:'
ws_hormon['A28'].font = SECTION_FONT
ws_hormon['A29'] = f'Aktuell: T:E2 = {round(np.mean(testosteron_werte)/np.mean(oestradiol_werte), 1)}:1'
ws_hormon['A30'] = 'Optimal: T:E2 = 15-20:1'

for col in ['A', 'B', 'C']:
    ws_hormon.column_dimensions[col].width = 35

# ============================================
# 5-7. DATENEINGABE (Früh, Mittag, Abend)
# ============================================
for name, data_arrays in [
    ('Früh', (libido_f, erektion_f, energie_f, stimmung_f, schlaf)),
    ('Mittag', (libido_m, erektion_m, energie_m, stimmung_m, None)),
    ('Abend', (libido_a, erektion_a, energie_a, stimmung_a, None))
]:
    ws_data = wb.create_sheet(f'Daten_{name}')
    
    ws_data['A1'] = f'{name.upper()} - Dateneingabe'
    ws_data['A1'].font = TITLE_FONT
    
    cols = ['Tag', 'Datum', 'Libido', 'Erektion', 'Energie', 'Stimmung']
    if name == 'Früh':
        cols.append('Schlaf')
    
    for i, h in enumerate(cols, 1):
        ws_data.cell(row=3, column=i).value = h
        ws_data.cell(row=3, column=i).fill = HEADER_FILL
        ws_data.cell(row=3, column=i).font = HEADER_FONT
    
    for day in range(n_days):
        row = 4 + day
        ws_data.cell(row=row, column=1).value = day + 1
        ws_data.cell(row=row, column=2).value = dates[day].strftime('%d.%m.%Y')
        ws_data.cell(row=row, column=3).value = data_arrays[0][day]
        ws_data.cell(row=row, column=4).value = data_arrays[1][day]
        ws_data.cell(row=row, column=5).value = data_arrays[2][day]
        ws_data.cell(row=row, column=6).value = data_arrays[3][day]
        if name == 'Früh':
            ws_data.cell(row=row, column=7).value = data_arrays[4][day]
    
    for col in ['A','B','C','D','E','F','G']:
        ws_data.column_dimensions[col].width = 12

# ============================================
# 8. HORMONWERTE EINGABE
# ============================================
ws_lab = wb.create_sheet('Hormonwerte')

ws_lab['A1'] = 'HORMONWERTE EINGABE'
ws_lab['A1'].font = TITLE_FONT

ws_lab['A3'] = 'Trage hier deine Laborwerte ein:'
ws_lab['A3'].font = Font(italic=True)

headers = ['Datum', 'Tag', 'Testosteron (ng/dL)', 'Östradiol (pg/mL)', 'Notizen']
for i, h in enumerate(headers, 1):
    ws_lab.cell(row=5, column=i).value = h
    ws_lab.cell(row=5, column=i).fill = HEADER_FILL
    ws_lab.cell(row=5, column=i).font = HEADER_FONT

for idx, tag in enumerate(mess_tage):
    row = 6 + idx
    ws_lab.cell(row=row, column=1).value = dates[tag].strftime('%d.%m.%Y')
    ws_lab.cell(row=row, column=2).value = tag + 1
    ws_lab.cell(row=row, column=3).value = testosteron_werte[idx]
    ws_lab.cell(row=row, column=4).value = oestradiol_werte[idx]

ws_lab['A11'] = 'REFERENZBEREICHE:'
ws_lab['A11'].font = SECTION_FONT
ws_lab['A12'] = 'Testosteron: 300-1000 ng/dL (optimal 500-800)'
ws_lab['A13'] = 'Östradiol: 10-40 pg/mL (optimal 20-30)'

for col in ['A','B','C','D','E']:
    ws_lab.column_dimensions[col].width = 20

# ============================================
# 9. ANLEITUNG
# ============================================
ws_help = wb.create_sheet('Anleitung')

ws_help['A1'] = 'ANLEITUNG'
ws_help['A1'].font = TITLE_FONT

anleitung = [
    '',
    'SO NUTZT DU DIESES DOKUMENT:',
    '',
    '1️⃣ TÄGLICH: Trage deine Werte ein',
    '   → Gehe zu "Daten_Früh", "Daten_Mittag", "Daten_Abend"',
    '   → Bewerte jeden Parameter von 0-10',
    '   → 0 = sehr schlecht, 10 = ausgezeichnet',
    '',
    '2️⃣ BEI BLUTABNAHME: Hormonwerte eintragen',
    '   → Gehe zu "Hormonwerte"',
    '   → Trage Testosteron und Östradiol ein',
    '',
    '3️⃣ AUSWERTUNG ANSEHEN:',
    '   → "AUSWERTUNG" = Dein automatischer Bericht',
    '   → "Verlaufsgrafiken" = Deine Verläufe im Detail',
    '   → "Hormon-Analyse" = Wie beeinflussen Hormone dich?',
    '',
    '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━',
    '',
    'WAS BEDEUTEN DIE KORRELATIONEN?',
    '',
    'r = +0.7 bis +1.0 → STARKER positiver Zusammenhang',
    'r = +0.3 bis +0.7 → MODERATER positiver Zusammenhang',
    'r = -0.3 bis +0.3 → KEIN deutlicher Zusammenhang',
    'r = -0.7 bis -0.3 → MODERATER negativer Zusammenhang',
    '',
    'Beispiel:',
    'Testosteron ↔ Erektion = r=0.8 bedeutet:',
    '"Wenn dein Testosteron höher ist, ist auch deine Erektion besser"',
]

for i, line in enumerate(anleitung, 2):
    ws_help.cell(row=i, column=1).value = line
    if line.startswith('SO NUTZT') or line.startswith('WAS BEDEUTEN'):
        ws_help.cell(row=i, column=1).font = SUBTITLE_FONT

ws_help.column_dimensions['A'].width = 60

# ============================================
# SPEICHERN
# ============================================
wb.save(OUTPUT_FILE)

print('✅ Datei erstellt:', OUTPUT_FILE)
print()
print('STRUKTUR:')
print('  📊 AUSWERTUNG        → Automatischer Bericht mit Status & Erkenntnissen')
print('  📉 Tagesdurchschnitt → Gesamtübersicht aller Parameter')
print('  📈 Verlaufsgrafiken  → Einzelne, übersichtliche Grafiken pro Parameter')
print('  🔬 Hormon-Analyse    → Testosteron & Östradiol Einfluss')
print('  📝 Daten_Früh/Mittag/Abend → Dateneingabe')
print('  💉 Hormonwerte       → Laborwerte eintragen')
print('  ❓ Anleitung         → Hilfe für den Nutzer')

#!/usr/bin/env python3
"""
Tagesprotokoll Analyzer
- Separiert Früh, Mittag, Abend Parameter
- Berechnet Korrelationen zwischen Hormonen (Testosteron, Östradiol) und anderen Parametern
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import os

# Datei-Pfade
INPUT_FILE = '/Users/lucasschmidt/Downloads/Tagesprotokoll_professionell_v5(2).xlsx'
OUTPUT_FILE = '/Users/lucasschmidt/Downloads/Tagesprotokoll_erweitert.xlsx'

def load_data():
    """Lädt die Eingabe- und Labor-Daten"""
    df_eingabe = pd.read_excel(INPUT_FILE, sheet_name='Eingabe')
    df_labor = pd.read_excel(INPUT_FILE, sheet_name='Labor')
    return df_eingabe, df_labor

def separate_by_zeitpunkt(df_eingabe):
    """Separiert die Daten nach Früh, Mittag, Abend"""
    # Filter für jeden Zeitpunkt
    df_frueh = df_eingabe[df_eingabe['Zeitpunkt'] == 'Früh'].copy()
    df_mittag = df_eingabe[df_eingabe['Zeitpunkt'] == 'Mittag'].copy()
    df_abend = df_eingabe[df_eingabe['Zeitpunkt'] == 'Abend'].copy()
    
    # Umbenennung der Spalten für Klarheit
    param_cols = ['Libido (0-10)', 'Erektionsqualität (0-10)', 'Energie (0-10)', 'Stimmung (0-10)', 'Schlafqualität (0-10)']
    
    for df, suffix in [(df_frueh, '_Früh'), (df_mittag, '_Mittag'), (df_abend, '_Abend')]:
        rename_dict = {}
        for col in param_cols:
            base_name = col.replace(' (0-10)', '')
            rename_dict[col] = base_name + suffix
        df.rename(columns=rename_dict, inplace=True)
        df.drop(columns=['Zeitpunkt', 'Notizen'], errors='ignore', inplace=True)
        df.reset_index(drop=True, inplace=True)
    
    return df_frueh, df_mittag, df_abend

def merge_with_labor(df_frueh, df_mittag, df_abend, df_labor):
    """Führt die Zeitpunkt-Daten mit Labor-Daten zusammen"""
    # Merge alle Zeitpunkte basierend auf Tag/Datum
    df_combined = df_frueh.merge(
        df_mittag.drop(columns=['Datum'], errors='ignore'), 
        on='Tag', 
        how='outer',
        suffixes=('', '_m')
    )
    df_combined = df_combined.merge(
        df_abend.drop(columns=['Datum'], errors='ignore'), 
        on='Tag', 
        how='outer',
        suffixes=('', '_a')
    )
    
    # Merge mit Labor-Daten
    df_combined = df_combined.merge(
        df_labor[['Tag', 'Testosteron', 'Östradiol']], 
        on='Tag', 
        how='outer'
    )
    
    df_combined.sort_values('Tag', inplace=True)
    df_combined.reset_index(drop=True, inplace=True)
    
    return df_combined

def calculate_correlations(df_combined):
    """Berechnet Korrelationen zwischen Hormonen und anderen Parametern"""
    
    # Parameter für jeden Zeitpunkt
    frueh_params = ['Libido_Früh', 'Erektionsqualität_Früh', 'Energie_Früh', 'Stimmung_Früh', 'Schlafqualität_Früh']
    mittag_params = ['Libido_Mittag', 'Erektionsqualität_Mittag', 'Energie_Mittag', 'Stimmung_Mittag']
    abend_params = ['Libido_Abend', 'Erektionsqualität_Abend', 'Energie_Abend', 'Stimmung_Abend']
    
    hormones = ['Testosteron', 'Östradiol']
    
    # Korrelationsmatrix erstellen
    correlation_results = []
    
    for hormone in hormones:
        row_frueh = {'Parameter': f'{hormone} vs Früh'}
        row_mittag = {'Parameter': f'{hormone} vs Mittag'}
        row_abend = {'Parameter': f'{hormone} vs Abend'}
        
        # Früh Korrelationen
        for param in frueh_params:
            if param in df_combined.columns and hormone in df_combined.columns:
                valid_data = df_combined[[hormone, param]].dropna()
                if len(valid_data) >= 3:
                    corr, p_value = stats.pearsonr(valid_data[hormone], valid_data[param])
                    param_short = param.replace('_Früh', '')
                    row_frueh[param_short] = corr
                    row_frueh[f'{param_short}_p'] = p_value
                else:
                    param_short = param.replace('_Früh', '')
                    row_frueh[param_short] = np.nan
                    row_frueh[f'{param_short}_p'] = np.nan
        
        # Mittag Korrelationen
        for param in mittag_params:
            if param in df_combined.columns and hormone in df_combined.columns:
                valid_data = df_combined[[hormone, param]].dropna()
                if len(valid_data) >= 3:
                    corr, p_value = stats.pearsonr(valid_data[hormone], valid_data[param])
                    param_short = param.replace('_Mittag', '')
                    row_mittag[param_short] = corr
                    row_mittag[f'{param_short}_p'] = p_value
                else:
                    param_short = param.replace('_Mittag', '')
                    row_mittag[param_short] = np.nan
                    row_mittag[f'{param_short}_p'] = np.nan
        
        # Abend Korrelationen
        for param in abend_params:
            if param in df_combined.columns and hormone in df_combined.columns:
                valid_data = df_combined[[hormone, param]].dropna()
                if len(valid_data) >= 3:
                    corr, p_value = stats.pearsonr(valid_data[hormone], valid_data[param])
                    param_short = param.replace('_Abend', '')
                    row_abend[param_short] = corr
                    row_abend[f'{param_short}_p'] = p_value
                else:
                    param_short = param.replace('_Abend', '')
                    row_abend[param_short] = np.nan
                    row_abend[f'{param_short}_p'] = np.nan
        
        correlation_results.append(row_frueh)
        correlation_results.append(row_mittag)
        correlation_results.append(row_abend)
    
    df_corr = pd.DataFrame(correlation_results)
    return df_corr

def create_detailed_correlation_table(df_combined):
    """Erstellt detaillierte Korrelationstabellen für Hormon-Einfluss"""
    
    params = ['Libido', 'Erektionsqualität', 'Energie', 'Stimmung', 'Schlafqualität']
    zeitpunkte = ['Früh', 'Mittag', 'Abend']
    hormones = ['Testosteron', 'Östradiol']
    
    results = []
    
    for hormone in hormones:
        for zeitpunkt in zeitpunkte:
            for param in params:
                col_name = f'{param}_{zeitpunkt}'
                if col_name in df_combined.columns and hormone in df_combined.columns:
                    valid_data = df_combined[[hormone, col_name]].dropna()
                    n = len(valid_data)
                    
                    if n >= 3:
                        corr, p_value = stats.pearsonr(valid_data[hormone], valid_data[col_name])
                        
                        # Interpretation der Korrelation
                        if abs(corr) < 0.1:
                            stärke = "keine"
                        elif abs(corr) < 0.3:
                            stärke = "schwach"
                        elif abs(corr) < 0.5:
                            stärke = "moderat"
                        elif abs(corr) < 0.7:
                            stärke = "stark"
                        else:
                            stärke = "sehr stark"
                        
                        richtung = "positiv" if corr > 0 else "negativ"
                        signifikant = "Ja" if p_value < 0.05 else "Nein"
                        
                        results.append({
                            'Hormon': hormone,
                            'Zeitpunkt': zeitpunkt,
                            'Parameter': param,
                            'Korrelation (r)': round(corr, 3),
                            'p-Wert': round(p_value, 4),
                            'Signifikant (p<0.05)': signifikant,
                            'Stärke': stärke,
                            'Richtung': richtung,
                            'N (Datenpunkte)': n
                        })
                    else:
                        results.append({
                            'Hormon': hormone,
                            'Zeitpunkt': zeitpunkt,
                            'Parameter': param,
                            'Korrelation (r)': np.nan,
                            'p-Wert': np.nan,
                            'Signifikant (p<0.05)': 'N/A',
                            'Stärke': 'zu wenig Daten',
                            'Richtung': 'N/A',
                            'N (Datenpunkte)': n
                        })
    
    return pd.DataFrame(results)

def create_summary_by_zeitpunkt(df_eingabe, df_labor):
    """Erstellt separate Übersichten für jeden Zeitpunkt"""
    
    summaries = {}
    
    for zeitpunkt in ['Früh', 'Mittag', 'Abend']:
        df_zeit = df_eingabe[df_eingabe['Zeitpunkt'] == zeitpunkt].copy()
        df_zeit = df_zeit.merge(df_labor[['Tag', 'Testosteron', 'Östradiol']], on='Tag', how='left')
        
        # Spalten umbenennen für Klarheit
        df_zeit = df_zeit.rename(columns={
            'Libido (0-10)': 'Libido',
            'Erektionsqualität (0-10)': 'Erektion',
            'Energie (0-10)': 'Energie',
            'Stimmung (0-10)': 'Stimmung',
            'Schlafqualität (0-10)': 'Schlaf'
        })
        
        # Nur relevante Spalten behalten
        cols = ['Tag', 'Datum', 'Libido', 'Erektion', 'Energie', 'Stimmung', 'Testosteron', 'Östradiol']
        if zeitpunkt == 'Früh':
            cols.insert(6, 'Schlaf')
        
        available_cols = [c for c in cols if c in df_zeit.columns]
        df_zeit = df_zeit[available_cols]
        df_zeit = df_zeit.drop(columns=['Zeitpunkt', 'Notizen'], errors='ignore')
        df_zeit.reset_index(drop=True, inplace=True)
        
        summaries[zeitpunkt] = df_zeit
    
    return summaries

def style_correlation_sheet(ws, df):
    """Formatiert das Korrelations-Sheet"""
    # Header Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Bedingte Formatierung für Korrelationswerte
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Farbskala für Korrelationen (rot = negativ, weiß = 0, grün = positiv)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Färbe Korrelationswerte
            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                if cell.column_letter == 'D':  # Korrelation (r) Spalte
                    if cell.value > 0.3:
                        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                    elif cell.value < -0.3:
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15

def create_output_file(df_eingabe, df_labor):
    """Erstellt die neue Excel-Datei mit allen Auswertungen"""
    
    # Separate Zeitpunkt-Übersichten
    summaries = create_summary_by_zeitpunkt(df_eingabe, df_labor)
    
    # Kombinierte Daten für Korrelationen
    df_frueh, df_mittag, df_abend = separate_by_zeitpunkt(df_eingabe)
    df_combined = merge_with_labor(df_frueh, df_mittag, df_abend, df_labor)
    
    # Detaillierte Korrelationstabelle
    df_corr_detail = create_detailed_correlation_table(df_combined)
    
    # Korrelationsmatrix (Testosteron)
    df_corr_testo = df_corr_detail[df_corr_detail['Hormon'] == 'Testosteron'].copy()
    
    # Korrelationsmatrix (Östradiol)
    df_corr_oestra = df_corr_detail[df_corr_detail['Hormon'] == 'Östradiol'].copy()
    
    # Excel-Writer
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # 1. Früh-Daten
        summaries['Früh'].to_excel(writer, sheet_name='Früh', index=False)
        
        # 2. Mittag-Daten
        summaries['Mittag'].to_excel(writer, sheet_name='Mittag', index=False)
        
        # 3. Abend-Daten
        summaries['Abend'].to_excel(writer, sheet_name='Abend', index=False)
        
        # 4. Kombinierte Daten
        df_combined.to_excel(writer, sheet_name='Alle Daten kombiniert', index=False)
        
        # 5. Testosteron-Korrelationen
        df_corr_testo.to_excel(writer, sheet_name='Korr_Testosteron', index=False)
        
        # 6. Östradiol-Korrelationen
        df_corr_oestra.to_excel(writer, sheet_name='Korr_Östradiol', index=False)
        
        # 7. Alle Korrelationen
        df_corr_detail.to_excel(writer, sheet_name='Alle Korrelationen', index=False)
    
    # Styling anwenden
    wb = load_workbook(OUTPUT_FILE)
    
    # Korrelations-Sheets formatieren
    for sheet_name in ['Korr_Testosteron', 'Korr_Östradiol', 'Alle Korrelationen']:
        ws = wb[sheet_name]
        df = df_corr_testo if 'Testo' in sheet_name else (df_corr_oestra if 'Östra' in sheet_name else df_corr_detail)
        style_correlation_sheet(ws, df)
    
    # Übersicht-Sheet erstellen
    ws_overview = wb.create_sheet('Übersicht', 0)
    
    # Übersicht-Inhalt
    overview_content = [
        ['TAGESPROTOKOLL - ERWEITERTE AUSWERTUNG', ''],
        ['', ''],
        ['SHEETS:', ''],
        ['Früh', 'Alle Parameter für den Zeitpunkt FRÜH mit Hormondaten'],
        ['Mittag', 'Alle Parameter für den Zeitpunkt MITTAG mit Hormondaten'],
        ['Abend', 'Alle Parameter für den Zeitpunkt ABEND mit Hormondaten'],
        ['Alle Daten kombiniert', 'Alle Zeitpunkte nebeneinander für Vergleiche'],
        ['', ''],
        ['KORRELATIONS-ANALYSE:', ''],
        ['Korr_Testosteron', 'Einfluss von Testosteron auf alle Parameter'],
        ['Korr_Östradiol', 'Einfluss von Östradiol auf alle Parameter'],
        ['Alle Korrelationen', 'Gesamtübersicht aller Hormon-Parameter-Korrelationen'],
        ['', ''],
        ['INTERPRETATION DER KORRELATIONEN:', ''],
        ['r > 0.7', 'Sehr starker positiver Zusammenhang'],
        ['0.5 < r < 0.7', 'Starker positiver Zusammenhang'],
        ['0.3 < r < 0.5', 'Moderater positiver Zusammenhang'],
        ['0.1 < r < 0.3', 'Schwacher positiver Zusammenhang'],
        ['-0.1 < r < 0.1', 'Kein Zusammenhang'],
        ['r < -0.3', 'Negativer Zusammenhang (höheres Hormon = niedrigerer Wert)'],
        ['', ''],
        ['SIGNIFIKANZ:', ''],
        ['p < 0.05', 'Statistisch signifikant (grün markiert)'],
        ['p >= 0.05', 'Nicht signifikant - könnte Zufall sein'],
    ]
    
    for row in overview_content:
        ws_overview.append(row)
    
    # Übersicht formatieren
    ws_overview['A1'].font = Font(bold=True, size=14)
    ws_overview.column_dimensions['A'].width = 25
    ws_overview.column_dimensions['B'].width = 50
    
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Datei erstellt: {OUTPUT_FILE}")
    print("\nEnthaltene Sheets:")
    print("  1. Übersicht - Erklärung aller Sheets")
    print("  2. Früh - Alle Parameter für FRÜH")
    print("  3. Mittag - Alle Parameter für MITTAG")
    print("  4. Abend - Alle Parameter für ABEND")
    print("  5. Alle Daten kombiniert - Für direkte Vergleiche")
    print("  6. Korr_Testosteron - Einfluss von Testosteron")
    print("  7. Korr_Östradiol - Einfluss von Östradiol")
    print("  8. Alle Korrelationen - Komplette Übersicht")

def main():
    print("🔄 Lade Daten...")
    df_eingabe, df_labor = load_data()
    
    print(f"   Eingabe-Daten: {len(df_eingabe)} Zeilen")
    print(f"   Labor-Daten: {len(df_labor)} Zeilen")
    
    print("\n🔄 Erstelle erweiterte Auswertung...")
    create_output_file(df_eingabe, df_labor)
    
    print("\n📊 Vorschau der Korrelationen:")
    df_frueh, df_mittag, df_abend = separate_by_zeitpunkt(df_eingabe)
    df_combined = merge_with_labor(df_frueh, df_mittag, df_abend, df_labor)
    df_corr = create_detailed_correlation_table(df_combined)
    
    # Zeige nur Zeilen mit genug Daten
    df_valid = df_corr[df_corr['N (Datenpunkte)'] >= 3]
    if not df_valid.empty:
        print(df_valid.to_string(index=False))
    else:
        print("   Noch nicht genug Datenpunkte für Korrelationsanalyse (min. 3 benötigt)")

if __name__ == "__main__":
    main()
